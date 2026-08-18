#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Script 2 â€” Itens das oportunidades nao convertidas (SEM ruÃ­do) + corte por RevisÃ£o do Gestor
LÃª:
- Arquivo gerado pelo Script 1 (aba: Lista_Oportunidades_Reais)
- Arquivo itens_do_orcamento com colunas:
  IDOrcamentoPrinc, DtInclusao, CodigoErp, Descricao, Quantidade, Valor (unitÃ¡rio), VlTot

Filtra nao convertidas (dentro das oportunidades reais, sem ruÃ­do):
- Nao convertidas = Faturou = 0 e Aprovado pelo Cliente = 0
E permite analisar:
- todas
- revisadas (Passou por RevisÃ£o Gestor = 1)
- sem_revisao (Passou por RevisÃ£o Gestor = 0)

Gera:
- Base_Nao_Convertidas
- Itens_Nao_Convertidas (linha a linha)
- Ranking_Itens_Por_Vendedor
- Ranking_Itens_Por_Mes
- Ranking_Itens_Vendedor_Mes
- Top10_<Vendedor> (uma aba por vendedor, top 10 por Valor_Nao_Convertido)
- Resumo_NC_Revisao (KPIs por tipo: todas/revisadas/sem revisÃ£o)
- Config

PreÃ§o mÃ©dio do item (gestÃ£o):
- Preco_Ponderado = Valor_Nao_Convertido / Volume_Nao_Convertido

Uso:
py -3.12 script2_itens_perdas_reais.py -i analise_codes_fev26_20260223_112350.xlsx -it itens_do_orcamento_fev26.xlsx -o itens_perdas_reais_fev26

Opcional:
--tipo_perda todas|revisadas|sem_revisao (padrao: todas)
--modo_data criacao|data_faturamento (padrao: criacao)
--start YYYY-MM-DD --end YYYY-MM-DD
--top 50
"""

import argparse
import os
import re
import time
import unicodedata
from datetime import date, datetime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


NAO_CONVERTIDA_STATUSES = [
    "Cancelado por Inatividade",
    "Orçamento Cancelado",
    "Em confecção",
]
RELEVANCIA_SHARE_MEDIA_PCT = 10
RELEVANCIA_SHARE_ALTA_PCT = 25
RELEVANCIA_RANK_MAX = 3
RELEVANCIA_SHARE_RANK_PCT = 5


# -----------------------------
# Args
# -----------------------------
def parse_args():
    def positive_int(value: str) -> int:
        ivalue = int(value)
        if ivalue <= 0:
            raise argparse.ArgumentTypeError("deve ser um inteiro maior que 0")
        return ivalue

    p = argparse.ArgumentParser()

    p.add_argument("-i", "--input", required=True, help="Arquivo gerado pelo Script 1 (.xlsx).")
    p.add_argument("-it", "--itens", required=True, help="Arquivo itens_do_orcamento (.xlsx/.xls/.csv).")
    p.add_argument("-o", "--output", required=True, help="Nome base do arquivo de saÃ­da (sem timestamp).")

    p.add_argument(
        "--tipo_perda",
        default="todas",
        choices=["todas", "revisadas", "sem_revisao"],
        help="Recorte de revisao: todas | revisadas | sem_revisao",
    )
    p.add_argument(
        "--modo_data",
        default="criacao",
        choices=["criacao", "data_faturamento", "ambos"],
        help="Define a data de referencia da analise: criacao | data_faturamento | ambos",
    )

    p.add_argument("--start", default=None, help="Data inÃ­cio (YYYY-MM-DD) para filtrar Data do orÃ§amento (opcional).")
    p.add_argument("--end", default=None, help="Data fim (YYYY-MM-DD) para filtrar Data do orÃ§amento (opcional).")
    p.add_argument("--top", type=positive_int, default=50, help="Top N itens para destacar no ranking geral.")
    p.add_argument(
        "--top_vendedor",
        type=positive_int,
        default=10,
        help="Top N itens por vendedor para abas 'Top10_<Vendedor>'.",
    )

    return p.parse_args()


# -----------------------------
# Utils
# -----------------------------
def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = (
        df.columns.astype(str)
        .str.replace("\ufeff", "", regex=False)
        .str.replace("\u200b", "", regex=False)
        .str.strip()
    )
    return df


def read_excel_sheet(path: str, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet_name)
    return clean_columns(df)


def apply_date_filter(df: pd.DataFrame, date_col: str, start: str | None, end: str | None) -> pd.DataFrame:
    out = df.copy()
    if start:
        dt_start = pd.to_datetime(start, errors="coerce")
        if pd.isna(dt_start):
            raise ValueError(f"Data invÃ¡lida em --start: '{start}'. Use YYYY-MM-DD.")
        out = out[out[date_col] >= dt_start]
    if end:
        dt_end = pd.to_datetime(end, errors="coerce")
        if pd.isna(dt_end):
            raise ValueError(f"Data invÃ¡lida em --end: '{end}'. Use YYYY-MM-DD.")
        dt_end_next = dt_end + pd.Timedelta(days=1)
        out = out[out[date_col] < dt_end_next]
    return out


def pick_first_existing(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for c in candidates:
        if c in df.columns:
            return c
    return None


def pick_status_orcamento_col(df: pd.DataFrame) -> str | None:
    return pick_first_existing(
        df,
        [
            "Status Atual do Orçamento",
            "Status Atual do Orcamento",
            "Status Atual do OrÃ§amento",
            "Status Atual",
            "Status Atual orçamento",
            "Status Atual orcamento",
            "Status Atual orÃ§amento",
            "Status Orçamento",
            "Status Orcamento",
            "Status OrÃ§amento",
            "Status",
        ],
    )


def normalize_status_text(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.lower().split())


def filter_nao_convertida_statuses(df: pd.DataFrame, status_col: str | None) -> pd.DataFrame:
    if not status_col or status_col not in df.columns:
        return df.iloc[0:0].copy()
    valid_statuses = {normalize_status_text(status) for status in NAO_CONVERTIDA_STATUSES}
    return df[df[status_col].map(normalize_status_text).isin(valid_statuses)].copy()


def join_orcamentos(values: pd.Series) -> str:
    ids = pd.to_numeric(values, errors="coerce").dropna().astype(int).unique().tolist()
    ids = sorted(ids)
    return ", ".join(str(x) for x in ids)


def add_price_reference_stats(ranking: pd.DataFrame, itens_base: pd.DataFrame, keys: list[str]) -> pd.DataFrame:
    if ranking.empty:
        ranking["Preco_Media_Orcamento"] = np.nan
        ranking["Preco_Mediana_Orcamento"] = np.nan
        return ranking

    if itens_base.empty:
        ranking["Preco_Media_Orcamento"] = np.nan
        ranking["Preco_Mediana_Orcamento"] = np.nan
        return ranking

    por_orcamento = (
        itens_base.groupby([*keys, "IDOrcamentoPrinc"], dropna=False)
        .agg(
            Volume_Orcamento=("Quantidade", "sum"),
            Valor_Orcamento=("VlTot", "sum"),
        )
        .reset_index()
    )
    por_orcamento["Preco_Por_Orcamento"] = np.where(
        por_orcamento["Volume_Orcamento"] > 0,
        por_orcamento["Valor_Orcamento"] / por_orcamento["Volume_Orcamento"],
        np.nan,
    )
    stats = (
        por_orcamento.groupby(keys, dropna=False)["Preco_Por_Orcamento"]
        .agg(
            Preco_Media_Orcamento="mean",
            Preco_Mediana_Orcamento="median",
        )
        .reset_index()
    )
    return ranking.merge(stats, on=keys, how="left")


def build_tipo_rankings(itens_base: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    keys = ["Tipo Cliente", "CodigoErp", "Descricao"]
    ranking = (
        itens_base.groupby(keys, dropna=False)
        .agg(
            Orcamentos_Nao_Convertidos=("IDOrcamentoPrinc", pd.Series.nunique),
            Linhas_Item=("IDOrcamentoPrinc", "count"),
            Volume_Nao_Convertido=("Quantidade", "sum"),
            Valor_Nao_Convertido=("VlTot", "sum"),
            Orcamentos_Lista=("IDOrcamentoPrinc", join_orcamentos),
        )
        .reset_index()
    )

    ranking["Preco_Ponderado"] = np.where(
        ranking["Volume_Nao_Convertido"] > 0,
        ranking["Valor_Nao_Convertido"] / ranking["Volume_Nao_Convertido"],
        np.nan,
    )
    ranking = add_price_reference_stats(ranking, itens_base, keys)

    totais_tipo = (
        itens_base.groupby("Tipo Cliente", dropna=False)
        .agg(
            Total_Orcamentos_Tipo=("IDOrcamentoPrinc", pd.Series.nunique),
            Total_Valor_Tipo=("VlTot", "sum"),
            Total_Volume_Tipo=("Quantidade", "sum"),
        )
        .reset_index()
    )
    ranking = ranking.merge(totais_tipo, on="Tipo Cliente", how="left")

    ranking["Share_Orcamentos_%"] = np.where(
        ranking["Total_Orcamentos_Tipo"] > 0,
        ranking["Orcamentos_Nao_Convertidos"] / ranking["Total_Orcamentos_Tipo"] * 100,
        np.nan,
    )
    ranking["Share_Valor_%"] = np.where(
        ranking["Total_Valor_Tipo"] > 0,
        ranking["Valor_Nao_Convertido"] / ranking["Total_Valor_Tipo"] * 100,
        np.nan,
    )
    ranking["Share_Volume_%"] = np.where(
        ranking["Total_Volume_Tipo"] > 0,
        ranking["Volume_Nao_Convertido"] / ranking["Total_Volume_Tipo"] * 100,
        np.nan,
    )

    ranking = ranking.drop(
        columns=["Total_Orcamentos_Tipo", "Total_Valor_Tipo", "Total_Volume_Tipo"]
    ).sort_values(["Tipo Cliente", "Valor_Nao_Convertido", "Volume_Nao_Convertido"], ascending=[True, False, False])

    tipo_norm = ranking["Tipo Cliente"].astype("string").fillna("").str.strip().str.lower()
    ranking_final = ranking[tipo_norm == "cliente final"].copy()
    ranking_revenda = ranking[tipo_norm == "revendedor"].copy()
    ranking_outros = ranking[~tipo_norm.isin(["cliente final", "revendedor"])].copy()

    return ranking, ranking_final, ranking_revenda, ranking_outros


def build_tipo_rankings_faturados(itens_base: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    keys = ["Tipo Cliente", "CodigoErp", "Descricao"]
    ranking = (
        itens_base.groupby(keys, dropna=False)
        .agg(
            Orcamentos_Faturados=("IDOrcamentoPrinc", pd.Series.nunique),
            Linhas_Item=("IDOrcamentoPrinc", "count"),
            Volume_Faturado=("Quantidade", "sum"),
            Valor_Faturado=("VlTot", "sum"),
            Orcamentos_Lista=("IDOrcamentoPrinc", join_orcamentos),
        )
        .reset_index()
    )

    ranking["Preco_Ponderado"] = np.where(
        ranking["Volume_Faturado"] > 0,
        ranking["Valor_Faturado"] / ranking["Volume_Faturado"],
        np.nan,
    )
    ranking = add_price_reference_stats(ranking, itens_base, keys)

    totais_tipo = (
        itens_base.groupby("Tipo Cliente", dropna=False)
        .agg(
            Total_Orcamentos_Tipo=("IDOrcamentoPrinc", pd.Series.nunique),
            Total_Valor_Tipo=("VlTot", "sum"),
            Total_Volume_Tipo=("Quantidade", "sum"),
        )
        .reset_index()
    )
    ranking = ranking.merge(totais_tipo, on="Tipo Cliente", how="left")

    ranking["Share_Orcamentos_%"] = np.where(
        ranking["Total_Orcamentos_Tipo"] > 0,
        ranking["Orcamentos_Faturados"] / ranking["Total_Orcamentos_Tipo"] * 100,
        np.nan,
    )
    ranking["Share_Valor_%"] = np.where(
        ranking["Total_Valor_Tipo"] > 0,
        ranking["Valor_Faturado"] / ranking["Total_Valor_Tipo"] * 100,
        np.nan,
    )
    ranking["Share_Volume_%"] = np.where(
        ranking["Total_Volume_Tipo"] > 0,
        ranking["Volume_Faturado"] / ranking["Total_Volume_Tipo"] * 100,
        np.nan,
    )

    ranking = ranking.drop(
        columns=["Total_Orcamentos_Tipo", "Total_Valor_Tipo", "Total_Volume_Tipo"]
    ).sort_values(["Tipo Cliente", "Valor_Faturado", "Volume_Faturado"], ascending=[True, False, False])

    tipo_norm = ranking["Tipo Cliente"].astype("string").fillna("").str.strip().str.lower()
    ranking_final = ranking[tipo_norm == "cliente final"].copy()
    ranking_revenda = ranking[tipo_norm == "revendedor"].copy()
    ranking_outros = ranking[~tipo_norm.isin(["cliente final", "revendedor"])].copy()

    return ranking, ranking_final, ranking_revenda, ranking_outros


def build_price_winloss_analysis(
    itens_faturados: pd.DataFrame,
    itens_nao_convertidos: pd.DataFrame,
    loss_label: str,
) -> pd.DataFrame:
    keys = ["Tipo Cliente", "CodigoErp", "Descricao"]

    def aggregate(df: pd.DataFrame, prefix: str) -> pd.DataFrame:
        id_list_col = f"Numeros_Orcamentos_{prefix}"
        if df.empty:
            return pd.DataFrame(
                columns=[
                    *keys,
                    f"Orcamentos_{prefix}",
                    id_list_col,
                    f"Linhas_{prefix}",
                    f"Volume_{prefix}",
                    f"Valor_{prefix}",
                    f"Preco_Ponderado_{prefix}",
                    f"Preco_Media_Orcamento_{prefix}",
                    f"Preco_Mediana_Orcamento_{prefix}",
                ]
            )

        def join_unique_ids(values: pd.Series) -> str:
            ids = values.dropna().astype(str).str.strip()
            ids = ids[ids != ""].drop_duplicates()
            return ", ".join(ids.sort_values().tolist())

        grouped = (
            df.groupby(keys, dropna=False)
            .agg(
                **{
                    f"Orcamentos_{prefix}": ("IDOrcamentoPrinc", pd.Series.nunique),
                    id_list_col: ("IDOrcamentoPrinc", join_unique_ids),
                    f"Linhas_{prefix}": ("IDOrcamentoPrinc", "count"),
                    f"Volume_{prefix}": ("Quantidade", "sum"),
                    f"Valor_{prefix}": ("VlTot", "sum"),
                }
            )
            .reset_index()
        )
        grouped[f"Preco_Ponderado_{prefix}"] = np.where(
            grouped[f"Volume_{prefix}"] > 0,
            grouped[f"Valor_{prefix}"] / grouped[f"Volume_{prefix}"],
            np.nan,
        )
        por_orcamento = (
            df.groupby([*keys, "IDOrcamentoPrinc"], dropna=False)
            .agg(
                Volume_Orcamento=("Quantidade", "sum"),
                Valor_Orcamento=("VlTot", "sum"),
            )
            .reset_index()
        )
        por_orcamento["Preco_Por_Orcamento"] = np.where(
            por_orcamento["Volume_Orcamento"] > 0,
            por_orcamento["Valor_Orcamento"] / por_orcamento["Volume_Orcamento"],
            np.nan,
        )
        stats = (
            por_orcamento.groupby(keys, dropna=False)["Preco_Por_Orcamento"]
            .agg(
                **{
                    f"Preco_Media_Orcamento_{prefix}": "mean",
                    f"Preco_Mediana_Orcamento_{prefix}": "median",
                }
            )
            .reset_index()
        )
        grouped = grouped.merge(stats, on=keys, how="left")
        return grouped

    wins = aggregate(itens_faturados, "Faturado").rename(
        columns={
            "Orcamentos_Faturado": "Orcamentos_Faturados",
            "Numeros_Orcamentos_Faturado": "Numeros_Orcamentos_Faturados",
            "Linhas_Faturado": "Linhas_Faturadas",
        }
    )
    losses = aggregate(itens_nao_convertidos, loss_label)
    if loss_label == "Nao_Convertido":
        losses = losses.rename(
            columns={
                "Orcamentos_Nao_Convertido": "Orcamentos_Nao_Convertidos",
                "Numeros_Orcamentos_Nao_Convertido": "Numeros_Orcamentos_Nao_Convertidos",
                "Linhas_Nao_Convertido": "Linhas_Nao_Convertidas",
            }
        )

    analysis = wins.merge(losses, on=keys, how="outer")
    id_list_cols = [c for c in analysis.columns if c.startswith("Numeros_Orcamentos_")]
    numeric_cols = [c for c in analysis.columns if c not in keys and c not in id_list_cols]
    for col in numeric_cols:
        analysis[col] = pd.to_numeric(analysis[col], errors="coerce")

    count_cols = [c for c in numeric_cols if c.startswith(("Orcamentos_", "Linhas_"))]
    volume_value_cols = [c for c in numeric_cols if c.startswith(("Volume_", "Valor_"))]
    analysis[count_cols] = analysis[count_cols].fillna(0).astype(int)
    analysis[volume_value_cols] = analysis[volume_value_cols].fillna(0.0)
    analysis[id_list_cols] = analysis[id_list_cols].fillna("")

    win_orc_col = "Orcamentos_Faturados"
    loss_orc_col = "Orcamentos_Nao_Convertidos" if loss_label == "Nao_Convertido" else f"Orcamentos_{loss_label}"
    loss_price_col = f"Preco_Ponderado_{loss_label}"
    loss_median_price_col = f"Preco_Mediana_Orcamento_{loss_label}"
    loss_value_col = f"Valor_{loss_label}"

    analysis["Total_Orcamentos_Comparados"] = analysis[win_orc_col] + analysis[loss_orc_col]
    analysis["Win_Rate_Item_%"] = np.where(
        analysis["Total_Orcamentos_Comparados"] > 0,
        analysis[win_orc_col] / analysis["Total_Orcamentos_Comparados"] * 100,
        np.nan,
    )
    analysis["Loss_Rate_Item_%"] = np.where(
        analysis["Total_Orcamentos_Comparados"] > 0,
        analysis[loss_orc_col] / analysis["Total_Orcamentos_Comparados"] * 100,
        np.nan,
    )
    analysis["Dif_Preco_%"] = np.where(
        (analysis["Preco_Ponderado_Faturado"] > 0) & (analysis[loss_price_col] > 0),
        (analysis[loss_price_col] / analysis["Preco_Ponderado_Faturado"] - 1) * 100,
        np.nan,
    )
    analysis["Dif_Preco_R$"] = np.where(
        analysis["Preco_Ponderado_Faturado"].notna() & analysis[loss_price_col].notna(),
        analysis[loss_price_col] - analysis["Preco_Ponderado_Faturado"],
        np.nan,
    )
    analysis["Dif_Preco_Mediana_%"] = np.where(
        (analysis["Preco_Mediana_Orcamento_Faturado"] > 0) & (analysis[loss_median_price_col] > 0),
        (analysis[loss_median_price_col] / analysis["Preco_Mediana_Orcamento_Faturado"] - 1) * 100,
        np.nan,
    )
    analysis["Dif_Preco_Mediana_R$"] = np.where(
        analysis["Preco_Mediana_Orcamento_Faturado"].notna() & analysis[loss_median_price_col].notna(),
        analysis[loss_median_price_col] - analysis["Preco_Mediana_Orcamento_Faturado"],
        np.nan,
    )

    price_threshold_pct = 3
    sample_ok = (analysis[win_orc_col] >= 3) & (analysis[loss_orc_col] >= 3)
    analysis["Classificacao_Preco"] = np.select(
        [
            ~sample_ok,
            (analysis["Dif_Preco_Mediana_%"] >= price_threshold_pct) & (analysis["Win_Rate_Item_%"] < 50),
            analysis["Dif_Preco_Mediana_%"] >= price_threshold_pct,
            analysis["Dif_Preco_Mediana_%"] <= -price_threshold_pct,
            analysis["Dif_Preco_Mediana_%"].abs() < price_threshold_pct,
        ],
        [
            "Amostra insuficiente",
            "Possivel sensibilidade a preco",
            "Atencao: preco nao convertido maior",
            "Nao convertido com preco menor",
            "Preco similar",
        ],
        default="Sem comparativo de preco",
    )

    return analysis.sort_values(
        ["Classificacao_Preco", loss_value_col, "Valor_Faturado"],
        ascending=[True, False, False],
    )


def add_item_relevance(itens_base: pd.DataFrame, origem_analise: str) -> pd.DataFrame:
    df = itens_base.copy()
    df["Origem_Analise"] = origem_analise

    if df.empty:
        df["Valor_Total_Orcamento_Itens"] = np.nan
        df["Share_Item_Orcamento_%"] = np.nan
        df["Rank_Item_No_Orcamento"] = np.nan
        df["Item_Relevante"] = "Nao"
        df["Classificacao_Relevancia"] = "Sem itens"
        df["Motivo_Relevancia"] = "Sem itens"
        return df

    df["Valor_Total_Orcamento_Itens"] = df.groupby("IDOrcamentoPrinc", dropna=False)["VlTot"].transform("sum")
    df["Share_Item_Orcamento_%"] = np.where(
        df["Valor_Total_Orcamento_Itens"] > 0,
        df["VlTot"] / df["Valor_Total_Orcamento_Itens"] * 100,
        np.nan,
    )
    df["Rank_Item_No_Orcamento"] = (
        df.groupby("IDOrcamentoPrinc", dropna=False)["VlTot"]
        .rank(method="first", ascending=False)
        .astype("Int64")
    )

    share = df["Share_Item_Orcamento_%"]
    rank = df["Rank_Item_No_Orcamento"]
    df["Classificacao_Relevancia"] = np.select(
        [
            share >= RELEVANCIA_SHARE_ALTA_PCT,
            share >= RELEVANCIA_SHARE_MEDIA_PCT,
            (rank <= RELEVANCIA_RANK_MAX) & (share >= RELEVANCIA_SHARE_RANK_PCT),
        ],
        [
            "Alta relevancia",
            "Media relevancia",
            "Relevante por rank",
        ],
        default="Baixa relevancia",
    )
    df["Item_Relevante"] = np.where(df["Classificacao_Relevancia"] == "Baixa relevancia", "Nao", "Sim")
    df["Motivo_Relevancia"] = df["Classificacao_Relevancia"]
    return df


def build_relevance_detail(itens_faturados: pd.DataFrame, itens_nao_convertidos: pd.DataFrame) -> pd.DataFrame:
    detalhe = pd.concat(
        [
            itens_faturados.copy(),
            itens_nao_convertidos.copy(),
        ],
        ignore_index=True,
        sort=False,
    )
    preferred_cols = [
        "Origem_Analise",
        "IDOrcamentoPrinc",
        "ID_Orcamento",
        "Data",
        "Tipo Cliente",
        "Vendedor",
        "CNPJ",
        "Status Atual",
        "Aprovado pelo Cliente",
        "Passou por Revisão Gestor",
        "CodigoErp",
        "Descricao",
        "Quantidade",
        "Valor",
        "VlTot",
        "Valor_Total_Orcamento_Itens",
        "Share_Item_Orcamento_%",
        "Rank_Item_No_Orcamento",
        "Item_Relevante",
        "Classificacao_Relevancia",
        "Motivo_Relevancia",
        "Periodo",
    ]
    cols = [c for c in preferred_cols if c in detalhe.columns]
    extra_cols = [c for c in detalhe.columns if c not in cols]
    return detalhe[cols + extra_cols]


def build_relevant_item_analysis(itens_faturados: pd.DataFrame, itens_nao_convertidos: pd.DataFrame) -> pd.DataFrame:
    relevantes_faturados = itens_faturados[itens_faturados["Item_Relevante"] == "Sim"].copy()
    relevantes_nao_convertidos = itens_nao_convertidos[itens_nao_convertidos["Item_Relevante"] == "Sim"].copy()

    analysis = build_price_winloss_analysis(relevantes_faturados, relevantes_nao_convertidos, "Nao_Convertido")
    if analysis.empty:
        return analysis

    keys = ["Tipo Cliente", "CodigoErp", "Descricao"]

    def relevance_stats(df: pd.DataFrame, prefix: str) -> pd.DataFrame:
        if df.empty:
            return pd.DataFrame(
                columns=[
                    *keys,
                    f"Share_Medio_Item_Orcamento_{prefix}_%",
                    f"Share_Mediano_Item_Orcamento_{prefix}_%",
                    f"Rank_Medio_Item_Orcamento_{prefix}",
                    f"Itens_Alta_Relevancia_{prefix}",
                    f"Itens_Media_Relevancia_{prefix}",
                    f"Itens_Relevantes_Por_Rank_{prefix}",
                ]
            )
        pivot_counts = (
            df.pivot_table(
                index=keys,
                columns="Classificacao_Relevancia",
                values="IDOrcamentoPrinc",
                aggfunc="count",
                fill_value=0,
                dropna=False,
            )
            .reset_index()
        )
        rename_counts = {
            "Alta relevancia": f"Itens_Alta_Relevancia_{prefix}",
            "Media relevancia": f"Itens_Media_Relevancia_{prefix}",
            "Relevante por rank": f"Itens_Relevantes_Por_Rank_{prefix}",
        }
        pivot_counts = pivot_counts.rename(columns=rename_counts)
        for col in rename_counts.values():
            if col not in pivot_counts.columns:
                pivot_counts[col] = 0

        stats = (
            df.groupby(keys, dropna=False)
            .agg(
                **{
                    f"Share_Medio_Item_Orcamento_{prefix}_%": ("Share_Item_Orcamento_%", "mean"),
                    f"Share_Mediano_Item_Orcamento_{prefix}_%": ("Share_Item_Orcamento_%", "median"),
                    f"Rank_Medio_Item_Orcamento_{prefix}": ("Rank_Item_No_Orcamento", "mean"),
                }
            )
            .reset_index()
        )
        return stats.merge(pivot_counts, on=keys, how="left")

    stats_fat = relevance_stats(relevantes_faturados, "Faturado")
    stats_nc = relevance_stats(relevantes_nao_convertidos, "Nao_Convertido")
    analysis = analysis.merge(stats_fat, on=keys, how="left").merge(stats_nc, on=keys, how="left")

    count_cols = [c for c in analysis.columns if c.startswith("Itens_")]
    analysis[count_cols] = analysis[count_cols].fillna(0).astype(int)
    return analysis.sort_values(
        ["Total_Orcamentos_Comparados", "Valor_Nao_Convertido", "Valor_Faturado"],
        ascending=[False, False, False],
    )


def read_itens(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext in [".xlsx", ".xls"]:
        df = pd.read_excel(
            path,
            dtype={
                "IDOrcamentoPrinc": "string",
                "CodigoErp": "string",
                "Descricao": "string",
            },
        )
        return clean_columns(df)

    if ext == ".csv":
        df = pd.read_csv(
            path,
            sep=None,
            engine="python",
            encoding="utf-8-sig",
            dtype={
                "IDOrcamentoPrinc": "string",
                "CodigoErp": "string",
                "Descricao": "string",
            },
        )
        return clean_columns(df)

    raise ValueError(f"ExtensÃ£o nÃ£o suportada: {ext}. Use .xlsx/.xls/.csv")


def build_output_path(output_stem: str) -> str:
    historico_dir = os.path.join(os.getcwd(), "historico")
    os.makedirs(historico_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{output_stem}_{timestamp}.xlsx"
    return os.path.join(historico_dir, filename)


def safe_sheet_name(name: str, max_len: int = 31) -> str:
    """
    Excel:
    - mÃ¡ximo 31 chars
    - nÃ£o pode: : \\ / ? * [ ]
    - evitar nomes vazios
    """
    s = str(name).strip()
    s = re.sub(r"[:\\/?*\[\]]", "_", s)
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        s = "SemNome"
    return s[:max_len]


def unique_sheet_name(base_name: str, used_names: set[str], max_len: int = 31) -> str:
    """
    Garante nome de aba Ãºnico dentro do arquivo.
    Se jÃ¡ existir, acrescenta sufixo _2, _3, ...
    """
    base = safe_sheet_name(base_name, max_len=max_len)
    candidate = base
    i = 2
    while candidate in used_names:
        suffix = f"_{i}"
        candidate = safe_sheet_name(base[: max_len - len(suffix)] + suffix, max_len=max_len)
        i += 1
    used_names.add(candidate)
    return candidate


def aplicar_formatacao_excel(path_xlsx: str):
    wb = load_workbook(path_xlsx)

    fmt_moeda = 'R$ #,##0.00'
    fmt_int = '#,##0'
    fmt_float = '#,##0'
    fmt_pct = '0.00"%"'
    fmt_data = "dd/mm/yyyy"

    for wsname in wb.sheetnames:
        ws = wb[wsname]
        if ws.max_row < 2:
            continue

        # Aplica quebra de texto em todos os tÃ­tulos (linha 1).
        # Mantemos a altura da linha sem valor fixo para o Excel ajustar ao abrir.
        for col in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col)
            if header_cell.value is not None:
                header_cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="center",
                    horizontal="left",
                )
        ws.row_dimensions[1].height = None

        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if not isinstance(header, str):
                continue
            h = header.lower()

            # percentuais (precisa vir antes de "valor", ex.: Share_Valor_%)
            if "share" in h or " %" in h or "percent" in h:
                fmt = fmt_pct
            # moeda
            elif "valor" in h or "vltot" in h or "preco" in h or "preÃ§o" in h:
                fmt = fmt_moeda
            # quantidades / volume
            elif "qtd" in h or "quant" in h or "volume" in h:
                fmt = fmt_float
            # contagens
            elif "orcamentos" in h or "orÃ§amentos" in h or "ocorr" in h or "linhas" in h:
                fmt = fmt_int
            else:
                continue

            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)) and not (isinstance(cell.value, float) and np.isnan(cell.value)):
                    cell.number_format = fmt

        # Datas em formato abreviado e CNPJ como texto.
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if isinstance(header, str) and header.strip().lower() == "cnpj":
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value is None:
                        continue
                    cell.value = str(cell.value)
                    cell.number_format = "@"

        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (datetime, date)):
                    cell.number_format = fmt_data

        # MantÃ©m textos alinhados Ã  esquerda em todas as abas.
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, str):
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical=cell.alignment.vertical if cell.alignment else None,
                        wrap_text=cell.alignment.wrap_text if cell.alignment else None,
                    )
                elif isinstance(cell.value, (int, float)) and not (isinstance(cell.value, float) and np.isnan(cell.value)):
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical=cell.alignment.vertical if cell.alignment else None,
                        wrap_text=cell.alignment.wrap_text if cell.alignment else None,
                    )
                elif isinstance(cell.value, (datetime, date)):
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical=cell.alignment.vertical if cell.alignment else None,
                        wrap_text=cell.alignment.wrap_text if cell.alignment else None,
                    )

        # Autoajuste de largura de coluna (com limite para evitar colunas excessivamente largas).
        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is None:
                    continue
                txt = str(val).strip()
                if txt:
                    max_len = max(max_len, len(txt))
            if max_len > 0:
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 2, 60)

    wb.save(path_xlsx)


def build_view_outputs(
    oportunidades: pd.DataFrame,
    itens: pd.DataFrame,
    args,
    col_status: str | None,
    modo_label: str,
) -> dict[str, pd.DataFrame | int | float]:
    faturados_totais = oportunidades[oportunidades["Faturou"] == 1].copy()
    faturados_revisados = faturados_totais[faturados_totais["Passou por Revisão Gestor"] == 1].copy()
    faturados_sem_revisao = faturados_totais[faturados_totais["Passou por Revisão Gestor"] == 0].copy()

    perdas_totais = oportunidades[
        (oportunidades["Faturou"] == 0) & (oportunidades["Aprovado pelo Cliente"] == 0)
    ].copy()
    perdas_totais = filter_nao_convertida_statuses(perdas_totais, col_status)
    perdas_revisadas = perdas_totais[perdas_totais["Passou por Revisão Gestor"] == 1].copy()
    perdas_sem_revisao = perdas_totais[perdas_totais["Passou por Revisão Gestor"] == 0].copy()

    perdas_totais = apply_date_filter(perdas_totais, "Data_Referencia_Analise", args.start, args.end)
    perdas_revisadas = apply_date_filter(perdas_revisadas, "Data_Referencia_Analise", args.start, args.end)
    perdas_sem_revisao = apply_date_filter(perdas_sem_revisao, "Data_Referencia_Analise", args.start, args.end)
    faturados_totais = apply_date_filter(faturados_totais, "Data_Referencia_Analise", args.start, args.end)
    faturados_revisados = apply_date_filter(faturados_revisados, "Data_Referencia_Analise", args.start, args.end)
    faturados_sem_revisao = apply_date_filter(faturados_sem_revisao, "Data_Referencia_Analise", args.start, args.end)

    if args.tipo_perda == "revisadas":
        perdas = perdas_revisadas
        faturados = faturados_revisados
    elif args.tipo_perda == "sem_revisao":
        perdas = perdas_sem_revisao
        faturados = faturados_sem_revisao
    else:
        perdas = perdas_totais
        faturados = faturados_totais

    perdas = perdas.dropna(subset=["ID_Orcamento"]).copy()
    perdas["ID_Orcamento"] = pd.to_numeric(perdas["ID_Orcamento"], errors="coerce").astype("Int64")
    perdas = perdas.dropna(subset=["ID_Orcamento"]).copy()
    perdas["ID_Orcamento"] = perdas["ID_Orcamento"].astype(int)

    faturados = faturados.dropna(subset=["ID_Orcamento"]).copy()
    faturados["ID_Orcamento"] = pd.to_numeric(faturados["ID_Orcamento"], errors="coerce").astype("Int64")
    faturados = faturados.dropna(subset=["ID_Orcamento"]).copy()
    faturados["ID_Orcamento"] = faturados["ID_Orcamento"].astype(int)

    perdas_ids = set(perdas["ID_Orcamento"].tolist())
    itens_perdas = itens[itens["IDOrcamentoPrinc"].isin(perdas_ids)].copy()
    faturados_ids = set(faturados["ID_Orcamento"].tolist())
    itens_faturados = itens[itens["IDOrcamentoPrinc"].isin(faturados_ids)].copy()

    perdas_ctx_cols = [
        "ID_Orcamento",
        "Data",
        "Data_Faturamento",
        "Data_Referencia_Analise",
        "Vendedor",
        "CNPJ",
        "Aprovado pelo Cliente",
        "Passou por Revisão Gestor",
        "Tipo Cliente",
    ]
    if col_status and col_status in perdas.columns:
        perdas_ctx_cols.append(col_status)
    perdas_ctx = perdas[perdas_ctx_cols].copy()
    itens_perdas = itens_perdas.merge(
        perdas_ctx,
        left_on="IDOrcamentoPrinc",
        right_on="ID_Orcamento",
        how="left",
    )

    faturados_ctx_cols = [
        "ID_Orcamento",
        "Data",
        "Data_Faturamento",
        "Data_Referencia_Analise",
        "Vendedor",
        "CNPJ",
        "Aprovado pelo Cliente",
        "Passou por Revisão Gestor",
        "Tipo Cliente",
    ]
    if col_status and col_status in faturados.columns:
        faturados_ctx_cols.append(col_status)
    faturados_ctx = faturados[faturados_ctx_cols].copy()
    itens_faturados = itens_faturados.merge(
        faturados_ctx,
        left_on="IDOrcamentoPrinc",
        right_on="ID_Orcamento",
        how="left",
    )

    itens_perdas["Periodo"] = (
        pd.to_datetime(itens_perdas["Data_Referencia_Analise"], errors="coerce").dt.to_period("M").dt.to_timestamp()
    )
    itens_faturados["Periodo"] = (
        pd.to_datetime(itens_faturados["Data_Referencia_Analise"], errors="coerce").dt.to_period("M").dt.to_timestamp()
    )
    itens_perdas = add_item_relevance(itens_perdas, "Nao Convertido")
    itens_faturados = add_item_relevance(itens_faturados, "Faturado")
    itens_relevancia_detalhe = build_relevance_detail(itens_faturados, itens_perdas)

    total_orc_perdidos = int(perdas["ID_Orcamento"].nunique())
    total_val_perdido_itens = float(itens_perdas["VlTot"].sum(skipna=True))
    total_vol_perdido = float(itens_perdas["Quantidade"].sum(skipna=True))

    ranking_geral = (
        itens_perdas.groupby(["CodigoErp", "Descricao"], dropna=False)
        .agg(
            Orcamentos_Nao_Convertidos=("IDOrcamentoPrinc", pd.Series.nunique),
            Linhas_Item=("IDOrcamentoPrinc", "count"),
            Volume_Nao_Convertido=("Quantidade", "sum"),
            Valor_Nao_Convertido=("VlTot", "sum"),
        )
        .reset_index()
    )
    ranking_geral["Preco_Ponderado"] = np.where(
        ranking_geral["Volume_Nao_Convertido"] > 0,
        ranking_geral["Valor_Nao_Convertido"] / ranking_geral["Volume_Nao_Convertido"],
        np.nan,
    )
    ranking_geral = add_price_reference_stats(ranking_geral, itens_perdas, ["CodigoErp", "Descricao"])
    ranking_geral["Share_Orcamentos_%"] = np.where(
        total_orc_perdidos > 0,
        ranking_geral["Orcamentos_Nao_Convertidos"] / total_orc_perdidos * 100,
        np.nan,
    )
    ranking_geral["Share_Valor_%"] = np.where(
        total_val_perdido_itens > 0,
        ranking_geral["Valor_Nao_Convertido"] / total_val_perdido_itens * 100,
        np.nan,
    )
    ranking_geral["Share_Volume_%"] = np.where(
        total_vol_perdido > 0,
        ranking_geral["Volume_Nao_Convertido"] / total_vol_perdido * 100,
        np.nan,
    )
    ranking_geral = ranking_geral.sort_values(
        ["Valor_Nao_Convertido", "Volume_Nao_Convertido", "Orcamentos_Nao_Convertidos"],
        ascending=False,
    )

    (
        ranking_geral_tipo,
        ranking_geral_tipo_final,
        ranking_geral_tipo_revenda,
        ranking_geral_tipo_outros,
    ) = build_tipo_rankings(itens_perdas)
    (
        ranking_faturados_tipo,
        ranking_faturados_tipo_final,
        ranking_faturados_tipo_revenda,
        ranking_faturados_tipo_outros,
    ) = build_tipo_rankings_faturados(itens_faturados)
    analise_preco_itens = build_price_winloss_analysis(
        itens_faturados,
        itens_perdas,
        "Nao_Convertido",
    )
    analise_itens_relevantes = build_relevant_item_analysis(
        itens_faturados,
        itens_perdas,
    )

    ranking_vendedor = (
        itens_perdas.groupby(["Vendedor", "CodigoErp", "Descricao"], dropna=False)
        .agg(
            Orcamentos_Nao_Convertidos=("IDOrcamentoPrinc", pd.Series.nunique),
            Linhas_Item=("IDOrcamentoPrinc", "count"),
            Volume_Nao_Convertido=("Quantidade", "sum"),
            Valor_Nao_Convertido=("VlTot", "sum"),
        )
        .reset_index()
    )
    ranking_vendedor["Preco_Ponderado"] = np.where(
        ranking_vendedor["Volume_Nao_Convertido"] > 0,
        ranking_vendedor["Valor_Nao_Convertido"] / ranking_vendedor["Volume_Nao_Convertido"],
        np.nan,
    )
    ranking_vendedor = add_price_reference_stats(ranking_vendedor, itens_perdas, ["Vendedor", "CodigoErp", "Descricao"])
    ranking_vendedor = ranking_vendedor.sort_values(["Vendedor", "Valor_Nao_Convertido"], ascending=[True, False])

    ranking_mes = (
        itens_perdas.groupby(["Periodo", "CodigoErp", "Descricao"], dropna=False)
        .agg(
            Orcamentos_Nao_Convertidos=("IDOrcamentoPrinc", pd.Series.nunique),
            Linhas_Item=("IDOrcamentoPrinc", "count"),
            Volume_Nao_Convertido=("Quantidade", "sum"),
            Valor_Nao_Convertido=("VlTot", "sum"),
        )
        .reset_index()
    )
    ranking_mes["Preco_Ponderado"] = np.where(
        ranking_mes["Volume_Nao_Convertido"] > 0,
        ranking_mes["Valor_Nao_Convertido"] / ranking_mes["Volume_Nao_Convertido"],
        np.nan,
    )
    ranking_mes = add_price_reference_stats(ranking_mes, itens_perdas, ["Periodo", "CodigoErp", "Descricao"])
    ranking_mes = ranking_mes.sort_values(["Periodo", "Valor_Nao_Convertido"], ascending=[True, False])

    ranking_vend_mes = (
        itens_perdas.groupby(["Periodo", "Vendedor", "CodigoErp", "Descricao"], dropna=False)
        .agg(
            Orcamentos_Nao_Convertidos=("IDOrcamentoPrinc", pd.Series.nunique),
            Linhas_Item=("IDOrcamentoPrinc", "count"),
            Volume_Nao_Convertido=("Quantidade", "sum"),
            Valor_Nao_Convertido=("VlTot", "sum"),
        )
        .reset_index()
    )
    ranking_vend_mes["Preco_Ponderado"] = np.where(
        ranking_vend_mes["Volume_Nao_Convertido"] > 0,
        ranking_vend_mes["Valor_Nao_Convertido"] / ranking_vend_mes["Volume_Nao_Convertido"],
        np.nan,
    )
    ranking_vend_mes = add_price_reference_stats(
        ranking_vend_mes,
        itens_perdas,
        ["Periodo", "Vendedor", "CodigoErp", "Descricao"],
    )
    ranking_vend_mes = ranking_vend_mes.sort_values(
        ["Periodo", "Vendedor", "Valor_Nao_Convertido"], ascending=[True, True, False]
    )

    def resumo_perdas(df_perdas: pd.DataFrame, label: str) -> dict:
        qtd = int(df_perdas["ID_Orcamento"].nunique()) if "ID_Orcamento" in df_perdas.columns else 0
        val = float(pd.to_numeric(df_perdas.get("Valor", pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
        return {"Tipo": label, "Qtd_Orcamentos_Nao_Convertidos": qtd, "Valor_Total_Nao_Convertido_(orcamento)": val}

    resumo = pd.DataFrame(
        [
            resumo_perdas(perdas_totais, "todas"),
            resumo_perdas(perdas_revisadas, "revisadas"),
            resumo_perdas(perdas_sem_revisao, "sem_revisao"),
        ]
    )

    config = pd.DataFrame(
        {
            "Parametro": [
                "input",
                "itens",
                "tipo_perda",
                "modo_data",
                "Filtro_Nao_Convertidas",
                "start",
                "end",
                "top",
                "top_vendedor",
                "Data_Execucao",
                "Qtd_Orcamentos_Nao_Convertidos_(recorte)",
                "Valor_Total_Nao_Convertido_Itens_(recorte)",
                "Volume_Total_Nao_Convertido_Itens_(recorte)",
                "Filtro_Status_Nao_Convertidas",
                "Criterio_Item_Relevante",
                "Share_Alta_Relevancia_%",
                "Share_Media_Relevancia_%",
                "Rank_Max_Relevante",
                "Share_Min_Rank_Relevante_%",
            ],
            "Valor": [
                args.input,
                args.itens,
                args.tipo_perda,
                modo_label,
                "Faturou = 0, Aprovado pelo Cliente = 0 e status em: " + ", ".join(NAO_CONVERTIDA_STATUSES),
                (args.start or ""),
                (args.end or ""),
                int(args.top),
                int(args.top_vendedor),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                total_orc_perdidos,
                total_val_perdido_itens,
                total_vol_perdido,
                ", ".join(NAO_CONVERTIDA_STATUSES),
                "Share >= 10% OU (Rank <= 3 E Share >= 5%)",
                RELEVANCIA_SHARE_ALTA_PCT,
                RELEVANCIA_SHARE_MEDIA_PCT,
                RELEVANCIA_RANK_MAX,
                RELEVANCIA_SHARE_RANK_PCT,
            ],
        }
    )

    return {
        "perdas": perdas,
        "itens_perdas": itens_perdas,
        "faturados": faturados,
        "itens_faturados": itens_faturados,
        "ranking_geral_tipo_final": ranking_geral_tipo_final,
        "ranking_geral_tipo_revenda": ranking_geral_tipo_revenda,
        "ranking_geral_tipo_outros": ranking_geral_tipo_outros,
        "ranking_faturados_tipo_final": ranking_faturados_tipo_final,
        "ranking_faturados_tipo_revenda": ranking_faturados_tipo_revenda,
        "ranking_faturados_tipo_outros": ranking_faturados_tipo_outros,
        "analise_preco_itens": analise_preco_itens,
        "itens_relevancia_detalhe": itens_relevancia_detalhe,
        "analise_itens_relevantes": analise_itens_relevantes,
        "ranking_vendedor": ranking_vendedor,
        "ranking_mes": ranking_mes,
        "ranking_vend_mes": ranking_vend_mes,
        "resumo": resumo,
        "config": config,
    }


# -----------------------------
# Main
# -----------------------------
def main():
    start_time = time.perf_counter()
    args = parse_args()

    # 1) Ler oportunidades reais (aba do Script 1)
    oportunidades = read_excel_sheet(args.input, "Lista_Oportunidades_Reais")

    col_id_orc = pick_first_existing(
        oportunidades,
        ["ID_Orcamento", "Núm. Orç.", "Num. Orc.", "NÃºm. OrÃ§.", "NÃƒÂºm. OrÃƒÂ§."],
    )
    col_data = pick_first_existing(
        oportunidades,
        ["Data", "Data de Criação", "Data de Criacao", "Data de CriaÃ§Ã£o", "Data de CriaÃƒÂ§ÃƒÂ£o"],
    )
    col_faturou = pick_first_existing(
        oportunidades,
        ["Faturou", "ETAPA 4 FUNIL Pedidos Faturados", "ETAPA 4 FUNIL Pedidos faturados"],
    )
    col_aprovado_cliente = pick_first_existing(
        oportunidades,
        [
            "Aprovado pelo Cliente",
            "Aprovados pelo Cliente",
            "ETAPA 3 FUNIL Aprovados pelo Cliente",
            "ETAPA 3 FUNIL Aprovado pelo Cliente",
            "ETAPA 3 FUNIL Aprovados pelo cliente",
            "ETAPA 3 FUNIL Aprovado pelo cliente",
        ],
    )
    col_revisao = pick_first_existing(
        oportunidades,
        [
            "Passou por Revisão Gestor",
            "Passou por revisão gestor",
            "Passou por Revisao Gestor",
            "Passou por RevisÃ£o Gestor",
            "Passou por revisÃ£o gestor",
            "Passou por RevisÃƒÂ£o Gestor",
        ],
    )

    required_cols = {
        "ID_Orcamento": col_id_orc,
        "Data": col_data,
        "Vendedor": "Vendedor" if "Vendedor" in oportunidades.columns else None,
        "CNPJ": "CNPJ" if "CNPJ" in oportunidades.columns else None,
        "Valor": "Valor" if "Valor" in oportunidades.columns else None,
        "Faturou": col_faturou,
        "Aprovado pelo Cliente": col_aprovado_cliente,
        "Passou por Revisão Gestor": col_revisao,
    }
    missing = [k for k, v in required_cols.items() if v is None]
    if missing:
        raise ValueError(
            "Colunas obrigatÃ³rias ausentes na aba Lista_Oportunidades_Reais:\n"
            + "\n".join(f"- {c}" for c in missing)
            + "\n\nColunas encontradas:\n"
            + ", ".join(oportunidades.columns.astype(str).tolist())
        )

    oportunidades = oportunidades.rename(
        columns={
            col_id_orc: "ID_Orcamento",
            col_data: "Data",
            col_faturou: "Faturou",
            col_aprovado_cliente: "Aprovado pelo Cliente",
            col_revisao: "Passou por Revisão Gestor",
        }
    )

    oportunidades["Data"] = pd.to_datetime(oportunidades["Data"], errors="coerce")
    col_data_faturamento = pick_first_existing(
        oportunidades,
        [
            "Data de Faturamento",
            "Data Faturamento",
            "Data_Faturamento",
            "Dt Faturamento",
            "Dt_Faturamento",
        ],
    )
    if col_data_faturamento and col_data_faturamento != "Data_Faturamento":
        oportunidades = oportunidades.rename(columns={col_data_faturamento: "Data_Faturamento"})
    if "Data_Faturamento" not in oportunidades.columns:
        oportunidades["Data_Faturamento"] = pd.NaT
    oportunidades["Data_Faturamento"] = pd.to_datetime(oportunidades["Data_Faturamento"], errors="coerce")
    col_data_ref_faturamento = pick_first_existing(
        oportunidades,
        [
            "Data Ref. Win Rate Faturamento",
            "Data Ref Win Rate Faturamento",
            "Data_Referencia_WR_Faturamento",
        ],
    )
    if col_data_ref_faturamento and col_data_ref_faturamento != "Data_Referencia_Analise_Faturamento":
        oportunidades = oportunidades.rename(
            columns={col_data_ref_faturamento: "Data_Referencia_Analise_Faturamento"}
        )
    if "Data_Referencia_Analise_Faturamento" not in oportunidades.columns:
        oportunidades["Data_Referencia_Analise_Faturamento"] = oportunidades["Data_Faturamento"].where(
            oportunidades["Data_Faturamento"].notna(),
            oportunidades["Data"],
        )
    oportunidades["Data_Referencia_Analise_Faturamento"] = pd.to_datetime(
        oportunidades["Data_Referencia_Analise_Faturamento"],
        errors="coerce",
    )
    if args.modo_data == "data_faturamento":
        oportunidades["Data_Referencia_Analise"] = oportunidades["Data_Referencia_Analise_Faturamento"]
    else:
        oportunidades["Data_Referencia_Analise"] = oportunidades["Data"]
    oportunidades["Faturou"] = pd.to_numeric(oportunidades["Faturou"], errors="coerce").fillna(0).astype(int)
    oportunidades["Aprovado pelo Cliente"] = (
        pd.to_numeric(oportunidades["Aprovado pelo Cliente"], errors="coerce").fillna(0).astype(int)
    )
    oportunidades["Valor"] = pd.to_numeric(oportunidades["Valor"], errors="coerce")

    oportunidades["Passou por Revisão Gestor"] = (
        pd.to_numeric(oportunidades["Passou por Revisão Gestor"], errors="coerce").fillna(0).astype(int)
    )
    if "Tipo Cliente" in oportunidades.columns:
        oportunidades["Tipo Cliente"] = oportunidades["Tipo Cliente"].astype("string").fillna("Não informado").str.strip()
        oportunidades.loc[oportunidades["Tipo Cliente"] == "", "Tipo Cliente"] = "Não informado"
    else:
        oportunidades["Tipo Cliente"] = "Não informado"

    col_status = pick_status_orcamento_col(oportunidades)
    itens = read_itens(args.itens)

    required_it = ["IDOrcamentoPrinc", "CodigoErp", "Descricao", "Quantidade", "Valor", "VlTot"]
    missing_it = [c for c in required_it if c not in itens.columns]
    if missing_it:
        raise ValueError(
            "Colunas obrigatÃ³rias ausentes em itens_do_orcamento:\n"
            + "\n".join(f"- {c}" for c in missing_it)
            + "\n\nColunas encontradas:\n"
            + ", ".join(itens.columns.astype(str).tolist())
        )

    itens["IDOrcamentoPrinc"] = pd.to_numeric(itens["IDOrcamentoPrinc"], errors="coerce").astype("Int64")
    itens = itens.dropna(subset=["IDOrcamentoPrinc"]).copy()
    itens = itens[itens["IDOrcamentoPrinc"] > 0].copy()
    itens["IDOrcamentoPrinc"] = itens["IDOrcamentoPrinc"].astype(int)

    itens["CodigoErp"] = itens["CodigoErp"].astype("string").fillna("").str.strip()
    itens["Descricao"] = itens["Descricao"].astype("string").fillna("").str.strip()

    itens["Quantidade"] = pd.to_numeric(itens["Quantidade"], errors="coerce").fillna(0.0)
    itens["Valor"] = pd.to_numeric(itens["Valor"], errors="coerce").fillna(0.0)   # unitÃ¡rio
    itens["VlTot"] = pd.to_numeric(itens["VlTot"], errors="coerce").fillna(0.0)   # total linha do item
    output_path = build_output_path(args.output)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        used_sheet_names = set(writer.book.sheetnames)
        if args.modo_data == "ambos":
            views = [
                ("criacao", "", oportunidades.assign(Data_Referencia_Analise=oportunidades["Data"])),
                (
                    "data_faturamento",
                    "_Data_Fat",
                    oportunidades.assign(
                        Data_Referencia_Analise=oportunidades["Data_Referencia_Analise_Faturamento"]
                    ),
                ),
            ]
        elif args.modo_data == "data_faturamento":
            views = [
                (
                    "data_faturamento",
                    "",
                    oportunidades.assign(Data_Referencia_Analise=oportunidades["Data_Referencia_Analise_Faturamento"]),
                )
            ]
        else:
            views = [("criacao", "", oportunidades.assign(Data_Referencia_Analise=oportunidades["Data"]))]

        top_vend_n = int(args.top_vendedor)

        for modo_label, suffix, oportunidades_view in views:
            view = build_view_outputs(
                oportunidades=oportunidades_view.copy(),
                itens=itens,
                args=args,
                col_status=col_status,
                modo_label=modo_label,
            )

            def sheet(base: str) -> str:
                return unique_sheet_name(f"{base}{suffix}", used_sheet_names)

            view["perdas"].to_excel(writer, sheet_name=sheet("Base_Nao_Convertidas"), index=False)
            view["itens_perdas"].to_excel(writer, sheet_name=sheet("Itens_Nao_Convertidas"), index=False)
            view["faturados"].to_excel(writer, sheet_name=sheet("Base_Faturados"), index=False)
            view["itens_faturados"].to_excel(writer, sheet_name=sheet("Itens_Faturados"), index=False)

            view["ranking_geral_tipo_final"].to_excel(writer, sheet_name=sheet("Ranking_Geral_Final"), index=False)
            view["ranking_geral_tipo_revenda"].to_excel(writer, sheet_name=sheet("Ranking_Geral_Revenda"), index=False)
            if not view["ranking_geral_tipo_outros"].empty:
                view["ranking_geral_tipo_outros"].to_excel(writer, sheet_name=sheet("Ranking_Geral_Outros"), index=False)
            view["ranking_faturados_tipo_final"].to_excel(
                writer,
                sheet_name=sheet("Ranking_Geral_Final_Faturados"),
                index=False,
            )
            view["ranking_faturados_tipo_revenda"].to_excel(
                writer,
                sheet_name=sheet("Ranking_Geral_Revenda_Faturados"),
                index=False,
            )
            if not view["ranking_faturados_tipo_outros"].empty:
                view["ranking_faturados_tipo_outros"].to_excel(
                    writer,
                    sheet_name=sheet("Ranking_Geral_Outros_Faturados"),
                    index=False,
                )
            view["analise_preco_itens"].to_excel(writer, sheet_name=sheet("Analise_Preco_Itens"), index=False)
            view["itens_relevancia_detalhe"].to_excel(
                writer,
                sheet_name=sheet("Itens_Relevancia_Detalhe"),
                index=False,
            )
            view["analise_itens_relevantes"].to_excel(
                writer,
                sheet_name=sheet("Analise_Itens_Relevantes"),
                index=False,
            )
            view["ranking_vendedor"].to_excel(writer, sheet_name=sheet("Ranking_Itens_Por_Vendedor"), index=False)
            view["ranking_mes"].to_excel(writer, sheet_name=sheet("Ranking_Itens_Por_Mes"), index=False)
            view["ranking_vend_mes"].to_excel(writer, sheet_name=sheet("Ranking_Itens_Vendedor_Mes"), index=False)
            view["resumo"].to_excel(writer, sheet_name=sheet("Resumo_NC_Revisao"), index=False)
            view["config"].to_excel(writer, sheet_name=sheet("Config"), index=False)

            vendedores = view["ranking_vendedor"]["Vendedor"].dropna().unique()
            for vend in vendedores:
                df_v = view["ranking_vendedor"][view["ranking_vendedor"]["Vendedor"] == vend].copy()

                total_val_vend = float(df_v["Valor_Nao_Convertido"].sum(skipna=True))
                total_vol_vend = float(df_v["Volume_Nao_Convertido"].sum(skipna=True))

                df_v["Share_Valor_%"] = np.where(
                    total_val_vend > 0,
                    df_v["Valor_Nao_Convertido"] / total_val_vend * 100,
                    np.nan,
                )
                df_v["Share_Volume_%"] = np.where(
                    total_vol_vend > 0,
                    df_v["Volume_Nao_Convertido"] / total_vol_vend * 100,
                    np.nan,
                )

                top_v = df_v.sort_values("Valor_Nao_Convertido", ascending=False).head(top_vend_n)
                top_sheet_base = f"Top{top_vend_n}_{vend}{suffix}"
                top_v.to_excel(writer, sheet_name=unique_sheet_name(top_sheet_base, used_sheet_names), index=False)

    aplicar_formatacao_excel(output_path)

    elapsed = time.perf_counter() - start_time
    elapsed_min = int(elapsed // 60)
    elapsed_sec = int(elapsed % 60)

    print(f"Arquivo gerado: {output_path}")
    print(f"Tempo total: {elapsed_min}min {elapsed_sec}s")


if __name__ == "__main__":
    main()
