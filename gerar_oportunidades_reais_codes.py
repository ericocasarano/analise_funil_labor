#!/usr/bin/env python
# -*- coding: utf-8 -*-

import argparse
import os
import time
import unicodedata
from datetime import date, datetime

import numpy as np
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Alignment


DEFINITIVE_LOSS_STATUSES = [
    "Cancelado por Inatividade",
    "Orçamento Cancelado",
    "Em confecção",
]


# -----------------------------
# Args
# -----------------------------
def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("-i", "--input", required=True, help="Arquivo base (remocao_ruidos) .xlsx/.xls/.csv")
    p.add_argument("-it", "--itens", required=True, help="Arquivo itens_do_orcamento .xlsx/.xls/.csv")
    p.add_argument("-o", "--output", required=True, help="Nome base do arquivo de saÃ­da (sem timestamp)")
    p.add_argument("--start", default="", help="Data inicial do recorte no formato YYYY-MM-DD")
    p.add_argument("--end", default="", help="Data final do recorte no formato YYYY-MM-DD")

    p.add_argument("--delta_horas", type=float, default=360, help="Janela em horas para agrupar no mesmo cluster")
    p.add_argument("--sim_min", type=float, default=0.50, help="Similaridade mÃ­nima (Jaccard) para cluster (0 a 1)")
    p.add_argument("--debug_id", type=int, default=None, help="ID para gerar aba Debug_ID_<ID> (opcional)")

    return p.parse_args()


# -----------------------------
# IO + NormalizaÃ§Ã£o
# -----------------------------
def read_input(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    if ext in [".xlsx", ".xls"]:
        # Aceita mÃºltiplos nomes possÃ­veis do documento como texto
        return pd.read_excel(
            path,
            dtype={
                "Cnpj": "string",
                "CNPJ": "string",
                "CNPJ/ CPF": "string",
                "CNPJ/CPF": "string",
            },
        )
    if ext == ".csv":
        return pd.read_csv(
            path,
            sep=None,
            engine="python",
            encoding="utf-8-sig",
            dtype={
                "Cnpj": "string",
                "CNPJ": "string",
                "CNPJ/ CPF": "string",
                "CNPJ/CPF": "string",
            },
        )
    raise ValueError(f"ExtensÃ£o nÃ£o suportada: {ext}. Use .xlsx/.xls/.csv")


def read_itens(path: str) -> pd.DataFrame:
    """
    LÃª IDOrcamentoPrinc como string para NÃƒO quebrar com linhas tipo "Total".
    Depois convertemos com errors='coerce'.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in [".xlsx", ".xls"]:
        return pd.read_excel(path, dtype={"IDOrcamentoPrinc": "string", "CodigoErp": "string"})
    if ext == ".csv":
        return pd.read_csv(
            path,
            sep=None,
            engine="python",
            encoding="utf-8-sig",
            dtype={"IDOrcamentoPrinc": "string", "CodigoErp": "string"},
        )
    raise ValueError(f"ExtensÃ£o nÃ£o suportada: {ext}. Use .xlsx/.xls/.csv")


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = (
        df.columns.astype(str)
        .str.replace("\ufeff", "", regex=False)
        .str.replace("\u200b", "", regex=False)
        .str.strip()
    )
    return df


def fix_mojibake_text(value: str) -> str:
    """
    Corrige textos com mojibake comum de UTF-8 lido como latin-1/cp1252.
    Ex.: "RuÃ­do" -> "Ruído".
    """
    if not isinstance(value, str):
        return value
    if "Ã" not in value and "Â" not in value:
        return value
    try:
        return value.encode("latin-1").decode("utf-8")
    except Exception:
        return value


def fix_df_headers_mojibake(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [fix_mojibake_text(str(c)) for c in df.columns]
    return df


def normalize_valor_to_numeric(series: pd.Series) -> pd.Series:
    def parse_one(v):
        if pd.isna(v):
            return np.nan
        if isinstance(v, (int, float, np.integer, np.floating)):
            return float(v)

        s = str(v).strip()
        if not s or s.lower() in {"nan", "none"}:
            return np.nan

        s = s.replace("\u00a0", " ")
        s = s.replace("R$", "").strip()
        s = s.replace(" ", "")

        if "," in s:
            s = s.replace(".", "").replace(",", ".")
            return pd.to_numeric(s, errors="coerce")

        if "." in s:
            parts = s.split(".")
            if len(parts[-1]) == 3 and all(p.isdigit() for p in parts):
                s = "".join(parts)

        return pd.to_numeric(s, errors="coerce")

    return series.apply(parse_one)


def normalize_doc_to_14(series: pd.Series) -> pd.Series:
    """
    MantÃ©m sÃ³ dÃ­gitos e garante 14 posiÃ§Ãµes (CNPJ).
    Se vier CPF (11), vira 14 com zeros Ã  esquerda (chave consistente).
    """
    s = series.astype("string")
    s = s.str.replace(r"\D", "", regex=True)
    s = s.fillna("")
    s = s.str.zfill(14)
    s = s.where(s != "00000000000000", "")
    return s


def pick_doc_col(df_cols) -> str:
    for possible in ["CNPJ/ CPF", "CNPJ/CPF", "CNPJ", "Cnpj"]:
        if possible in df_cols:
            return possible
    return ""


def pick_first_existing(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for c in candidates:
        if c in df.columns:
            return c
    return None


def pick_status_orcamento_col(df: pd.DataFrame) -> str | None:
    # 1) tenta nomes exatos mais comuns
    exact = pick_first_existing(
        df,
        [
            "Status Atual do OrÃ§amento",
            "Status Atual do Orcamento",
            "Status Atual",
            "Status Atual orÃ§amento",
            "Status Atual orcamento",
            "Status OrÃ§amento",
            "Status Orcamento",
            "Status",
        ],
    )
    if exact:
        return exact

    # 2) fallback: procura qualquer coluna contendo "status" + "orc"
    for c in df.columns.astype(str).tolist():
        c_norm = (
            c.lower()
            .replace("Ã§", "c")
            .replace("Ã£", "a")
            .replace("Ã¡", "a")
            .replace("Ã¢", "a")
            .replace("Ã©", "e")
            .replace("Ãª", "e")
            .replace("Ã­", "i")
            .replace("Ã³", "o")
            .replace("Ã´", "o")
            .replace("Ãµ", "o")
            .replace("Ãº", "u")
        )
        if "status" in c_norm and "orc" in c_norm:
            return c
    return None


def normalize_status_text(value) -> str:
    if pd.isna(value):
        return ""
    text = fix_mojibake_text(str(value)).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return " ".join(text.lower().split())


def filter_definitive_losses(df: pd.DataFrame, status_col: str | None) -> pd.DataFrame:
    if not status_col or status_col not in df.columns:
        return df.iloc[0:0].copy()
    definitive_statuses = {normalize_status_text(status) for status in DEFINITIVE_LOSS_STATUSES}
    return df[df[status_col].map(normalize_status_text).isin(definitive_statuses)].copy()


# -----------------------------
# Similaridade (Jaccard por CÃ³digo)
# -----------------------------
def jaccard_codes(a: set, b: set) -> float:
    if not a or not b:
        return 0.0
    inter = len(a & b)
    uni = len(a | b)
    return inter / uni if uni > 0 else 0.0


# -----------------------------
# Output
# -----------------------------
def build_output_path(output_stem: str) -> str:
    historico_dir = os.path.join(os.getcwd(), "historico")
    os.makedirs(historico_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{output_stem}_{timestamp}.xlsx"
    return os.path.join(historico_dir, filename)


def apply_date_filter(df: pd.DataFrame, date_col: str, start: str = "", end: str = "") -> pd.DataFrame:
    out = df.copy()

    if start:
        start_dt = pd.to_datetime(start, errors="coerce")
        if pd.isna(start_dt):
            raise ValueError(f"Data inicial invalida: {start}. Use o formato YYYY-MM-DD.")
        out = out[out[date_col] >= start_dt].copy()

    if end:
        end_dt = pd.to_datetime(end, errors="coerce")
        if pd.isna(end_dt):
            raise ValueError(f"Data final invalida: {end}. Use o formato YYYY-MM-DD.")
        out = out[out[date_col] <= end_dt].copy()

    return out


def build_clustered_funil(df: pd.DataFrame, mapa_codes: dict[int, set], delta_horas: float, sim_min: float):
    base = df[df["Enviado_Aprovacao"] == 1].copy()
    base = base.sort_values(["Vendedor", "CNPJ", "Data", "ID_Orcamento"]).reset_index(drop=True)

    base["DeltaHoras"] = base.groupby(["Vendedor", "CNPJ"])["Data"].diff().dt.total_seconds() / 3600
    base["Prev_ID"] = base.groupby(["Vendedor", "CNPJ"])["ID_Orcamento"].shift(1)

    sims = []
    for cur, prev in zip(base["ID_Orcamento"].tolist(), base["Prev_ID"].tolist()):
        if pd.isna(prev):
            sims.append(np.nan)
            continue
        try:
            cur_i = int(cur)
            prev_i = int(prev)
        except Exception:
            sims.append(np.nan)
            continue

        sims.append(
            jaccard_codes(
                mapa_codes.get(cur_i, set()),
                mapa_codes.get(prev_i, set()),
            )
        )
    base["Itens_Similarity"] = sims

    base["Crit_Tempo_OK"] = base["DeltaHoras"].notna() & (base["DeltaHoras"] <= delta_horas)
    base["Crit_Itens_OK"] = base["Itens_Similarity"].fillna(0.0) >= sim_min
    base["Relacionado"] = base["Crit_Tempo_OK"] & base["Crit_Itens_OK"]

    base["Novo_Cluster"] = (~base["Relacionado"]).astype(int)
    base.loc[base.groupby(["Vendedor", "CNPJ"]).cumcount() == 0, "Novo_Cluster"] = 1
    base["Cluster_ID"] = base.groupby(["Vendedor", "CNPJ"])["Novo_Cluster"].cumsum()

    base["Oportunidade_Real"] = 0
    base["Ruido"] = 0
    base["Motivo_Ruido"] = ""
    base["Cluster_Tem_Faturado"] = 0
    base["ID_Ultimo_Cluster"] = np.nan

    for (_, _, _), grp in base.groupby(["Vendedor", "CNPJ", "Cluster_ID"], sort=False):
        tem_faturado = int((grp["Faturou"] == 1).any())
        last_id = grp.iloc[-1]["ID_Orcamento"]

        base.loc[grp.index, "Cluster_Tem_Faturado"] = tem_faturado
        base.loc[grp.index, "ID_Ultimo_Cluster"] = last_id

        # Um cluster e' cortado em segmentos a cada faturado: tudo que vem
        # ANTES de um faturado, dentro do mesmo segmento, foi superado por ele
        # (recotacao da mesma negociacao) e vira ruido. Um orcamento criado
        # DEPOIS de um faturado inicia um segmento novo - nao e' mais a mesma
        # negociacao (que ja fechou), entao e' avaliado de forma independente.
        segmento = []
        for idx, faturou in zip(grp.index, grp["Faturou"]):
            segmento.append(idx)
            if faturou == 1:
                base.loc[idx, "Oportunidade_Real"] = 1
                anteriores = segmento[:-1]
                if anteriores:
                    base.loc[anteriores, "Ruido"] = 1
                    base.loc[anteriores, "Motivo_Ruido"] = "Superado por faturado posterior no mesmo cluster"
                segmento = []

        if segmento:
            last_idx = segmento[-1]
            base.loc[last_idx, "Oportunidade_Real"] = 1
            anteriores = segmento[:-1]
            if anteriores:
                base.loc[anteriores, "Ruido"] = 1
                base.loc[anteriores, "Motivo_Ruido"] = "Sem faturado no cluster e nÃ£o Ã© o Ãºltimo do cluster"

    base["Valor_Faturado"] = np.where(base["Faturou"] == 1, base["Valor"], 0.0)
    base["Valor_Ruido"] = np.where(base["Ruido"] == 1, base["Valor"], 0.0)
    base["Data_Referencia_WR_Faturamento"] = base["Data_Faturamento"].where(
        base["Data_Faturamento"].notna(),
        base["Data"],
    )

    oportunidades = base[base["Oportunidade_Real"] == 1].copy()
    ruidos = base[base["Ruido"] == 1].copy()
    return base, oportunidades, ruidos


# -----------------------------
# FormataÃ§Ã£o Excel (Comparativos)
# -----------------------------
def aplicar_formatacao_excel(path_xlsx: str):
    wb = load_workbook(path_xlsx)

    fmt_moeda = 'R$ #,##0'
    fmt_pct = '0.00"%"'  # win rate em 0..100
    fmt_data = "dd/mm/yyyy"
    money_headers = {
        "valor_ruido",
        "valor_faturado",
        "valor",
        "valor_total_nao_convertido",
        "valor_orcamentos_passou_revisao",
        "valor_orcamentos_sem_revisao",
    }
    percent_headers = {"share_qtd_orcamentos_%", "share_valor_%"}

    # Aplica quebra de texto no tÃ­tulo (linha 1) de todas as abas.
    for wsname in wb.sheetnames:
        ws = wb[wsname]
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            if cell.value is not None:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="center",
                    horizontal=cell.alignment.horizontal if cell.alignment else None,
                )
        ws.row_dimensions[1].height = None

    # Abas consolidadas por mÃ©trica na coluna A
    for sheet_name in ["Comparativo_Geral_Total", "Comp_Geral_Total_Data_Fat"]:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            metrica = ws.cell(row=row, column=1).value
            if not isinstance(metrica, str):
                continue
            m = metrica.lower()

            if "win rate" in m:
                fmt = fmt_pct
            elif "valor" in m:
                fmt = fmt_moeda
            else:
                continue

            for col in (2, 3):
                cell = ws.cell(row=row, column=col)

                if isinstance(cell.value, str):
                    s = cell.value.replace("R$", "").replace("%", "").replace(" ", "")
                    s = s.replace(".", "").replace(",", ".")
                    try:
                        cell.value = float(s)
                    except Exception:
                        continue

                if isinstance(cell.value, (int, float)):
                    cell.number_format = fmt

    # Garante formatação monetária nas colunas específicas em todas as abas.
    for wsname in wb.sheetnames:
        ws = wb[wsname]
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if not isinstance(header, str):
                continue
            h = header.strip().lower()
            if h in money_headers:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell.value, (int, float)) and not (isinstance(cell.value, float) and np.isnan(cell.value)):
                        cell.number_format = fmt_moeda
            elif h in percent_headers:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell.value, (int, float)) and not (isinstance(cell.value, float) and np.isnan(cell.value)):
                        cell.number_format = fmt_pct

    # Alinhamento base em todas as abas: texto à esquerda e número centralizado.
    for wsname in wb.sheetnames:
        ws = wb[wsname]
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, str):
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical=cell.alignment.vertical if cell.alignment else None,
                        wrap_text=cell.alignment.wrap_text if cell.alignment else None,
                    )
                elif isinstance(cell.value, (datetime, date)):
                    cell.number_format = fmt_data
                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical=cell.alignment.vertical if cell.alignment else None,
                        wrap_text=cell.alignment.wrap_text if cell.alignment else None,
                    )
                elif isinstance(cell.value, (int, float)) and not (isinstance(cell.value, float) and np.isnan(cell.value)):
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical=cell.alignment.vertical if cell.alignment else None,
                        wrap_text=cell.alignment.wrap_text if cell.alignment else None,
                    )

    # Autoajuste de largura de coluna em todas as abas.
    for wsname in wb.sheetnames:
        ws = wb[wsname]
        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is None:
                    continue
                txt = str(val).strip()
                if txt:
                    max_len = max(max_len, len(txt))
            if max_len > 0:
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 2, 60)

    wb.save(path_xlsx)


# -----------------------------
# Main
# -----------------------------
def main():
    start_time = time.perf_counter()
    args = parse_args()

    # -------- base --------
    df_raw = clean_columns(read_input(args.input))

    col_doc = pick_doc_col(df_raw.columns)
    if not col_doc:
        raise ValueError("Coluna de documento nÃ£o encontrada. Esperado: Cnpj / CNPJ / CNPJ/ CPF / CNPJ/CPF")

    col_id_orc = pick_first_existing(df_raw, ["Núm. Orç.", "Num. Orc.", "NÃºm. OrÃ§."])
    col_data_criacao = pick_first_existing(df_raw, ["Data de Criação", "Data de Criacao", "Data de CriaÃ§Ã£o"])
    col_data_faturamento = pick_first_existing(
        df_raw,
        [
            "Data de Faturamento",
            "Data Faturamento",
            "Data_Faturamento",
            "Dt Faturamento",
            "Dt_Faturamento",
        ],
    )
    col_etapa2 = pick_first_existing(
        df_raw,
        [
            "ETAPA 2 FUNIL Orçamentos em Negociação",
            "ETAPA 2 FUNIL Orcamentos em Negociacao",
            "ETAPA 2 FUNIL OrÃ§amentos em NegociaÃ§Ã£o",
        ],
    )
    col_etapa4 = pick_first_existing(
        df_raw,
        [
            "ETAPA 4 FUNIL Pedidos Faturados",
            "ETAPA 4 FUNIL Pedidos faturados",
        ],
    )

    required_base = {
        "ID_Orcamento": col_id_orc,
        "Data": col_data_criacao,
        "CNPJ": col_doc,
        "Vendedor": "Vendedor" if "Vendedor" in df_raw.columns else None,
        "Valor": "Valor" if "Valor" in df_raw.columns else None,
        "Enviado_Aprovacao": col_etapa2,
        "Faturou": col_etapa4,
    }
    missing = [k for k, v in required_base.items() if v is None]
    if missing:
        raise ValueError(
            "Colunas obrigatÃ³rias ausentes em remocao_ruidos:\n"
            + "\n".join(f"- {c}" for c in missing)
            + "\n\nColunas encontradas:\n"
            + ", ".join(df_raw.columns.astype(str).tolist())
        )

    # Renomeia internamente para padronizar o restante do fluxo
    rename_map = {
        col_id_orc: "ID_Orcamento",
        col_data_criacao: "Data",
        col_doc: "CNPJ",
        "Vendedor": "Vendedor",
        "Valor": "Valor",
        col_etapa2: "Enviado_Aprovacao",
        col_etapa4: "Faturou",
    }
    if col_data_faturamento:
        rename_map[col_data_faturamento] = "Data_Faturamento"

    df = df_raw.rename(columns=rename_map).copy()

    df["Data"] = pd.to_datetime(df["Data"], errors="coerce")
    if "Data_Faturamento" not in df.columns:
        df["Data_Faturamento"] = pd.NaT
    df["Data_Faturamento"] = pd.to_datetime(df["Data_Faturamento"], errors="coerce")
    df["Valor"] = normalize_valor_to_numeric(df["Valor"])
    df["Faturou"] = pd.to_numeric(df["Faturou"], errors="coerce").fillna(0).astype(int)
    df["Enviado_Aprovacao"] = pd.to_numeric(df["Enviado_Aprovacao"], errors="coerce").fillna(0).astype(int)

    # Documento como texto normalizado (14 dÃ­gitos)
    df["CNPJ"] = normalize_doc_to_14(df["CNPJ"])

    # -------- itens --------
    itens_raw = clean_columns(read_itens(args.itens))
    required_itens = ["IDOrcamentoPrinc", "CodigoErp"]
    missing_it = [c for c in required_itens if c not in itens_raw.columns]
    if missing_it:
        raise ValueError(
            "Colunas obrigatÃ³rias ausentes em itens_do_orcamento:\n"
            + "\n".join(f"- {c}" for c in missing_it)
            + "\n\nColunas encontradas:\n"
            + ", ".join(itens_raw.columns.astype(str).tolist())
        )

    itens = itens_raw.copy()
    itens["IDOrcamentoPrinc"] = pd.to_numeric(itens["IDOrcamentoPrinc"], errors="coerce").astype("Int64")
    itens["CodigoErp"] = itens["CodigoErp"].astype("string").fillna("").str.strip()

    # remove linhas invÃ¡lidas (ex: Total)
    itens = itens.dropna(subset=["IDOrcamentoPrinc"])
    itens = itens[itens["IDOrcamentoPrinc"] > 0]

    mapa_codes = {
        int(oid): set(sub["CodigoErp"].tolist())
        for oid, sub in itens.groupby("IDOrcamentoPrinc")
    }

    # A remocao de ruido roda uma unica vez sobre o historico completo (nao so
    # sobre a fatia do periodo), para que uma recotacao seja reconhecida mesmo
    # quando a negociacao original ficou fora da janela analisada. As duas
    # visoes (Data de Criacao / Data de Faturamento) diferem apenas em qual
    # coluna de data seleciona quem pertence ao periodo atual.
    base_full, oportunidades_full, ruidos_full = build_clustered_funil(
        df,
        mapa_codes,
        args.delta_horas,
        args.sim_min,
    )

    base = apply_date_filter(base_full, "Data", args.start, args.end)
    oportunidades = apply_date_filter(oportunidades_full, "Data", args.start, args.end)
    ruidos = apply_date_filter(ruidos_full, "Data", args.start, args.end)
    perdas_reais = oportunidades[oportunidades["Faturou"] == 0].copy()

    base_ref_periodo = apply_date_filter(
        base_full,
        "Data_Referencia_WR_Faturamento",
        args.start,
        args.end,
    )
    oportunidades_ref_periodo = apply_date_filter(
        oportunidades_full,
        "Data_Referencia_WR_Faturamento",
        args.start,
        args.end,
    )
    oportunidades_ref_periodo = oportunidades_ref_periodo.copy()

    # Aba auxiliar: perdas reais com foco em status/qtd/valor (quando disponÃ­veis)
    col_status = pick_status_orcamento_col(perdas_reais)
    col_qtd_pedidos = pick_first_existing(
        perdas_reais,
        [
            "Quantidade de Pedidos",
            "Qtd Pedidos",
            "Qtde Pedidos",
            "Qtd_Pedidos",
            "Quantidade Pedidos",
        ],
    )
    col_revisao_gestor = pick_first_existing(
        perdas_reais,
        [
            "Passou por Revisão Gestor",
            "Passou por revisão gestor",
            "Passou por RevisÃ£o Gestor",
            "Passou por Revisao Gestor",
            "Revisão Gestor",
            "revisão gestor",
            "RevisÃ£o Gestor",
            "Revisao Gestor",
        ],
    )

    perdas_reais_definitivas = filter_definitive_losses(perdas_reais, col_status)

    cols_perdas = ["ID_Orcamento", "Data", "Vendedor", "CNPJ"]
    if col_revisao_gestor:
        cols_perdas.append(col_revisao_gestor)
    if col_status:
        cols_perdas.append(col_status)
    if col_qtd_pedidos:
        cols_perdas.append(col_qtd_pedidos)
    cols_perdas += ["Valor", "Faturou"]

    # Remove duplicadas caso algum nome jÃ¡ exista na lista-base
    cols_perdas = [c for i, c in enumerate(cols_perdas) if c in perdas_reais.columns and c not in cols_perdas[:i]]
    perdas_reais_resumo = perdas_reais[cols_perdas].copy()
    perdas_reais_definitivas_resumo = perdas_reais_definitivas[cols_perdas].copy()

    # Pivot / tabela dinamica de oportunidades nao convertidas por Status Atual
    if col_status:
        pivot_perdas_status = (
            perdas_reais.groupby(col_status, dropna=False)
            .agg(
                Qtd_Orcamentos_Nao_Convertidos=("ID_Orcamento", pd.Series.nunique),
                Linhas=("ID_Orcamento", "count"),
                Valor_Total_Nao_Convertido=("Valor", "sum"),
            )
            .reset_index()
            .rename(columns={col_status: "Status Atual"})
        )

        if col_qtd_pedidos and col_qtd_pedidos in perdas_reais.columns:
            pedidos_por_status = (
                perdas_reais.groupby(col_status, dropna=False)[col_qtd_pedidos]
                .sum(min_count=1)
                .reset_index()
                .rename(columns={col_status: "Status Atual", col_qtd_pedidos: "Qtd_Pedidos_Total"})
            )
            pivot_perdas_status = pivot_perdas_status.merge(pedidos_por_status, on="Status Atual", how="left")

        if col_revisao_gestor and col_revisao_gestor in perdas_reais.columns:
            perdas_reais_aux = perdas_reais.assign(
                _rev_num=pd.to_numeric(perdas_reais[col_revisao_gestor], errors="coerce").fillna(0).astype(int),
                _valor_num=pd.to_numeric(perdas_reais["Valor"], errors="coerce").fillna(0),
            )
            rev_por_status = (
                perdas_reais_aux.groupby(col_status, dropna=False)
                .agg(
                    Qtd_Orcamentos_Passou_Revisao=(
                        "ID_Orcamento",
                        lambda s: int(s[perdas_reais_aux.loc[s.index, "_rev_num"] == 1].nunique()),
                    ),
                    Qtd_Orcamentos_Sem_Revisao=(
                        "ID_Orcamento",
                        lambda s: int(s[perdas_reais_aux.loc[s.index, "_rev_num"] == 0].nunique()),
                    ),
                    Valor_Orcamentos_Passou_Revisao=(
                        "_valor_num",
                        lambda s: float(s[perdas_reais_aux.loc[s.index, "_rev_num"] == 1].sum()),
                    ),
                    Valor_Orcamentos_Sem_Revisao=(
                        "_valor_num",
                        lambda s: float(s[perdas_reais_aux.loc[s.index, "_rev_num"] == 0].sum()),
                    ),
                )
                .reset_index()
                .rename(columns={col_status: "Status Atual"})
            )
            pivot_perdas_status = pivot_perdas_status.merge(rev_por_status, on="Status Atual", how="left")

        total_qtd_perdas = float(pivot_perdas_status["Qtd_Orcamentos_Nao_Convertidos"].sum())
        total_val_perdas = float(pd.to_numeric(pivot_perdas_status["Valor_Total_Nao_Convertido"], errors="coerce").fillna(0).sum())

        pivot_perdas_status["Share_Qtd_Orcamentos_%"] = np.where(
            total_qtd_perdas > 0,
            pivot_perdas_status["Qtd_Orcamentos_Nao_Convertidos"] / total_qtd_perdas * 100,
            np.nan,
        )
        pivot_perdas_status["Share_Valor_%"] = np.where(
            total_val_perdas > 0,
            pd.to_numeric(pivot_perdas_status["Valor_Total_Nao_Convertido"], errors="coerce").fillna(0) / total_val_perdas * 100,
            np.nan,
        )

        pivot_perdas_status = pivot_perdas_status.sort_values(
            ["Valor_Total_Nao_Convertido", "Qtd_Orcamentos_Nao_Convertidos"],
            ascending=False,
        )

        # Ordena e seleciona colunas da Pivot conforme padrao solicitado.
        pivot_cols_order = [
            "Status Atual",
            "Valor_Total_Nao_Convertido",
            "Qtd_Orcamentos_Passou_Revisao",
            "Valor_Orcamentos_Passou_Revisao",
            "Qtd_Orcamentos_Sem_Revisao",
            "Valor_Orcamentos_Sem_Revisao",
            "Share_Qtd_Orcamentos_%",
            "Share_Valor_%",
        ]
        pivot_perdas_status = pivot_perdas_status[[c for c in pivot_cols_order if c in pivot_perdas_status.columns]]
    else:
        pivot_perdas_status = pd.DataFrame(
            {
                "Aviso": [
                    "Coluna de Status Atual do orÃ§amento nÃ£o encontrada na base. "
                    "Aba Lista_Nao_Convertidas foi gerada sem segmentaÃ§Ã£o por status."
                ]
            }
        )

    if col_status:
        pivot_perdas_status_definitivas = (
            perdas_reais_definitivas.groupby(col_status, dropna=False)
            .agg(
                Qtd_Orcamentos_Nao_Convertidos=("ID_Orcamento", pd.Series.nunique),
                Linhas=("ID_Orcamento", "count"),
                Valor_Total_Nao_Convertido=("Valor", "sum"),
            )
            .reset_index()
            .rename(columns={col_status: "Status Atual"})
        )

        if col_qtd_pedidos and col_qtd_pedidos in perdas_reais_definitivas.columns:
            pedidos_def_por_status = (
                perdas_reais_definitivas.groupby(col_status, dropna=False)[col_qtd_pedidos]
                .sum(min_count=1)
                .reset_index()
                .rename(columns={col_status: "Status Atual", col_qtd_pedidos: "Qtd_Pedidos_Total"})
            )
            pivot_perdas_status_definitivas = pivot_perdas_status_definitivas.merge(
                pedidos_def_por_status,
                on="Status Atual",
                how="left",
            )

        if col_revisao_gestor and col_revisao_gestor in perdas_reais_definitivas.columns:
            perdas_def_aux = perdas_reais_definitivas.assign(
                _rev_num=pd.to_numeric(perdas_reais_definitivas[col_revisao_gestor], errors="coerce").fillna(0).astype(int),
                _valor_num=pd.to_numeric(perdas_reais_definitivas["Valor"], errors="coerce").fillna(0),
            )
            rev_def_por_status = (
                perdas_def_aux.groupby(col_status, dropna=False)
                .agg(
                    Qtd_Orcamentos_Passou_Revisao=(
                        "ID_Orcamento",
                        lambda s: int(s[perdas_def_aux.loc[s.index, "_rev_num"] == 1].nunique()),
                    ),
                    Qtd_Orcamentos_Sem_Revisao=(
                        "ID_Orcamento",
                        lambda s: int(s[perdas_def_aux.loc[s.index, "_rev_num"] == 0].nunique()),
                    ),
                    Valor_Orcamentos_Passou_Revisao=(
                        "_valor_num",
                        lambda s: float(s[perdas_def_aux.loc[s.index, "_rev_num"] == 1].sum()),
                    ),
                    Valor_Orcamentos_Sem_Revisao=(
                        "_valor_num",
                        lambda s: float(s[perdas_def_aux.loc[s.index, "_rev_num"] == 0].sum()),
                    ),
                )
                .reset_index()
                .rename(columns={col_status: "Status Atual"})
            )
            pivot_perdas_status_definitivas = pivot_perdas_status_definitivas.merge(
                rev_def_por_status,
                on="Status Atual",
                how="left",
            )

        total_qtd_perdas_def = float(pivot_perdas_status_definitivas["Qtd_Orcamentos_Nao_Convertidos"].sum())
        total_val_perdas_def = float(
            pd.to_numeric(pivot_perdas_status_definitivas["Valor_Total_Nao_Convertido"], errors="coerce").fillna(0).sum()
        )

        pivot_perdas_status_definitivas["Share_Qtd_Orcamentos_%"] = np.where(
            total_qtd_perdas_def > 0,
            pivot_perdas_status_definitivas["Qtd_Orcamentos_Nao_Convertidos"] / total_qtd_perdas_def * 100,
            np.nan,
        )
        pivot_perdas_status_definitivas["Share_Valor_%"] = np.where(
            total_val_perdas_def > 0,
            pd.to_numeric(pivot_perdas_status_definitivas["Valor_Total_Nao_Convertido"], errors="coerce").fillna(0)
            / total_val_perdas_def
            * 100,
            np.nan,
        )

        pivot_perdas_status_definitivas = pivot_perdas_status_definitivas.sort_values(
            ["Valor_Total_Nao_Convertido", "Qtd_Orcamentos_Nao_Convertidos"],
            ascending=False,
        )
        pivot_cols_order = [
            "Status Atual",
            "Valor_Total_Nao_Convertido",
            "Qtd_Orcamentos_Passou_Revisao",
            "Valor_Orcamentos_Passou_Revisao",
            "Qtd_Orcamentos_Sem_Revisao",
            "Valor_Orcamentos_Sem_Revisao",
            "Share_Qtd_Orcamentos_%",
            "Share_Valor_%",
        ]
        pivot_perdas_status_definitivas = pivot_perdas_status_definitivas[
            [c for c in pivot_cols_order if c in pivot_perdas_status_definitivas.columns]
        ]
    else:
        pivot_perdas_status_definitivas = pd.DataFrame(
            {
                "Aviso": [
                    "Coluna de Status Atual do orÃ§amento nÃ£o encontrada na base. "
                    "Aba Lista_NC_Recusa_Cliente foi gerada sem aplicar o filtro de status."
                ]
            }
        )

    # -----------------------------
    # Comparativo_Geral_Total (consolidado)
    # -----------------------------
    bruto_qtd_enviados = int(base["ID_Orcamento"].count())
    bruto_qtd_faturado = int(base["Faturou"].sum())
    bruto_qtd_perdas = bruto_qtd_enviados - bruto_qtd_faturado

    bruto_val_enviado = float(base["Valor"].sum(skipna=True))
    bruto_val_faturado = float(base["Valor_Faturado"].sum(skipna=True))
    bruto_val_perdas = bruto_val_enviado - bruto_val_faturado

    bruto_wr_vol = (bruto_qtd_faturado / bruto_qtd_enviados * 100) if bruto_qtd_enviados > 0 else np.nan
    bruto_wr_val = (bruto_val_faturado / bruto_val_enviado * 100) if bruto_val_enviado > 0 else np.nan

    limpo_qtd_enviados = int(oportunidades["ID_Orcamento"].count())
    limpo_qtd_faturado = int(oportunidades["Faturou"].sum())
    limpo_qtd_perdas = limpo_qtd_enviados - limpo_qtd_faturado

    limpo_val_enviado = float(oportunidades["Valor"].sum(skipna=True))
    limpo_val_faturado = float(oportunidades["Valor_Faturado"].sum(skipna=True))
    limpo_val_perdas = limpo_val_enviado - limpo_val_faturado

    limpo_wr_vol = (limpo_qtd_faturado / limpo_qtd_enviados * 100) if limpo_qtd_enviados > 0 else np.nan
    limpo_wr_val = (limpo_val_faturado / limpo_val_enviado * 100) if limpo_val_enviado > 0 else np.nan

    bruto_ref_qtd_enviados = int(base_ref_periodo["ID_Orcamento"].count())
    bruto_ref_qtd_faturado = int(base_ref_periodo["Faturou"].sum())
    bruto_ref_val_enviado = float(base_ref_periodo["Valor"].sum(skipna=True))
    bruto_ref_val_faturado = float(base_ref_periodo["Valor_Faturado"].sum(skipna=True))
    bruto_wr_vol_ref = (
        bruto_ref_qtd_faturado / bruto_ref_qtd_enviados * 100
        if bruto_ref_qtd_enviados > 0 else np.nan
    )
    bruto_wr_val_ref = (
        bruto_ref_val_faturado / bruto_ref_val_enviado * 100
        if bruto_ref_val_enviado > 0 else np.nan
    )

    limpo_ref_qtd_enviados = int(oportunidades_ref_periodo["ID_Orcamento"].count())
    limpo_ref_qtd_faturado = int(oportunidades_ref_periodo["Faturou"].sum())
    limpo_ref_val_enviado = float(oportunidades_ref_periodo["Valor"].sum(skipna=True))
    limpo_ref_val_faturado = float(oportunidades_ref_periodo["Valor_Faturado"].sum(skipna=True))
    limpo_wr_vol_ref = (
        limpo_ref_qtd_faturado / limpo_ref_qtd_enviados * 100
        if limpo_ref_qtd_enviados > 0 else np.nan
    )
    limpo_wr_val_ref = (
        limpo_ref_val_faturado / limpo_ref_val_enviado * 100
        if limpo_ref_val_enviado > 0 else np.nan
    )

    comparativo_total = pd.DataFrame(
        [
            ["Qtd Enviados", bruto_qtd_enviados, limpo_qtd_enviados],
            ["Qtd Faturados", bruto_qtd_faturado, limpo_qtd_faturado],
            ["Nao Convertidas (Qtd)", bruto_qtd_perdas, limpo_qtd_perdas],
            ["Valor Enviado", bruto_val_enviado, limpo_val_enviado],
            ["Valor Faturado", bruto_val_faturado, limpo_val_faturado],
            ["Nao Convertidas (Valor)", bruto_val_perdas, limpo_val_perdas],
            ["Win Rate (Volume) %", bruto_wr_vol, limpo_wr_vol],
            ["Win Rate (Valor) %", bruto_wr_val, limpo_wr_val],
        ],
        columns=["MÃ©trica", "Com RuÃ­do", "Sem RuÃ­do"],
    )

    comparativo_total_data_fat = pd.DataFrame(
        [
            ["Qtd Enviados", bruto_ref_qtd_enviados, limpo_ref_qtd_enviados],
            ["Qtd Faturados", bruto_ref_qtd_faturado, limpo_ref_qtd_faturado],
            [
                "Nao Convertidas (Qtd)",
                bruto_ref_qtd_enviados - bruto_ref_qtd_faturado,
                limpo_ref_qtd_enviados - limpo_ref_qtd_faturado,
            ],
            ["Valor Enviado", bruto_ref_val_enviado, limpo_ref_val_enviado],
            ["Valor Faturado", bruto_ref_val_faturado, limpo_ref_val_faturado],
            [
                "Nao Convertidas (Valor)",
                bruto_ref_val_enviado - bruto_ref_val_faturado,
                limpo_ref_val_enviado - limpo_ref_val_faturado,
            ],
            ["Win Rate (Volume) %", bruto_wr_vol_ref, limpo_wr_vol_ref],
            ["Win Rate (Valor) %", bruto_wr_val_ref, limpo_wr_val_ref],
        ],
        columns=["MÃ©trica", "Com RuÃ­do", "Sem RuÃ­do"],
    )

    # -----------------------------
    # Comparativo_Geral (MENSAL)
    # -----------------------------
    base["Periodo"] = base["Data"].dt.to_period("M").dt.to_timestamp()
    oportunidades["Periodo"] = oportunidades["Data"].dt.to_period("M").dt.to_timestamp()

    bruto_mes = (
        base.groupby("Periodo", dropna=False)
        .agg(
            Enviados_Qtd=("ID_Orcamento", "count"),
            Enviados_Valor=("Valor", "sum"),
            Faturado_Qtd=("Faturou", "sum"),
            Faturado_Valor=("Valor_Faturado", "sum"),
        )
        .reset_index()
    )
    bruto_mes["Perdas_Qtd"] = bruto_mes["Enviados_Qtd"] - bruto_mes["Faturado_Qtd"]
    bruto_mes["Perdas_Valor"] = bruto_mes["Enviados_Valor"] - bruto_mes["Faturado_Valor"]
    bruto_mes["Win Rate (Volume) %"] = np.where(
        bruto_mes["Enviados_Qtd"] > 0,
        bruto_mes["Faturado_Qtd"] / bruto_mes["Enviados_Qtd"] * 100,
        np.nan,
    )
    bruto_mes["Win Rate (Valor) %"] = np.where(
        bruto_mes["Enviados_Valor"] > 0,
        bruto_mes["Faturado_Valor"] / bruto_mes["Enviados_Valor"] * 100,
        np.nan,
    )

    limpo_mes = (
        oportunidades.groupby("Periodo", dropna=False)
        .agg(
            Enviados_Qtd=("ID_Orcamento", "count"),
            Enviados_Valor=("Valor", "sum"),
            Faturado_Qtd=("Faturou", "sum"),
            Faturado_Valor=("Valor_Faturado", "sum"),
        )
        .reset_index()
    )
    limpo_mes["Perdas_Qtd"] = limpo_mes["Enviados_Qtd"] - limpo_mes["Faturado_Qtd"]
    limpo_mes["Perdas_Valor"] = limpo_mes["Enviados_Valor"] - limpo_mes["Faturado_Valor"]
    limpo_mes["Win Rate (Volume) %"] = np.where(
        limpo_mes["Enviados_Qtd"] > 0,
        limpo_mes["Faturado_Qtd"] / limpo_mes["Enviados_Qtd"] * 100,
        np.nan,
    )
    limpo_mes["Win Rate (Valor) %"] = np.where(
        limpo_mes["Enviados_Valor"] > 0,
        limpo_mes["Faturado_Valor"] / limpo_mes["Enviados_Valor"] * 100,
        np.nan,
    )

    comparativo_mensal = bruto_mes.merge(
        limpo_mes,
        on="Periodo",
        how="outer",
        suffixes=(" (Com RuÃ­do)", " (Sem RuÃ­do)"),
    ).sort_values("Periodo")

    comparativo_mensal = comparativo_mensal[
        [
            "Periodo",
            "Enviados_Qtd (Com RuÃ­do)",
            "Enviados_Valor (Com RuÃ­do)",
            "Faturado_Qtd (Com RuÃ­do)",
            "Faturado_Valor (Com RuÃ­do)",
            "Perdas_Qtd (Com RuÃ­do)",
            "Perdas_Valor (Com RuÃ­do)",
            "Win Rate (Volume) % (Com RuÃ­do)",
            "Win Rate (Valor) % (Com RuÃ­do)",
            "Enviados_Qtd (Sem RuÃ­do)",
            "Enviados_Valor (Sem RuÃ­do)",
            "Faturado_Qtd (Sem RuÃ­do)",
            "Faturado_Valor (Sem RuÃ­do)",
            "Perdas_Qtd (Sem RuÃ­do)",
            "Perdas_Valor (Sem RuÃ­do)",
            "Win Rate (Volume) % (Sem RuÃ­do)",
            "Win Rate (Valor) % (Sem RuÃ­do)",
        ]
    ]

    # -----------------------------
    # Config_Modelo
    # -----------------------------
    config = pd.DataFrame(
        {
            "Parametro": ["delta_horas", "sim_min", "debug_id", "Data_Execucao", "Coluna_Documento"],
            "Valor": [
                args.delta_horas,
                args.sim_min,
                ("" if args.debug_id is None else args.debug_id),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                col_doc,
            ],
        }
    )

    # -----------------------------
    # Debug_Clusters
    # -----------------------------
    debug_cols = [
        "ID_Orcamento", "Data", "Vendedor", "CNPJ",
        "Prev_ID", "DeltaHoras", "Itens_Similarity",
        "Crit_Tempo_OK", "Crit_Itens_OK", "Relacionado",
        "Novo_Cluster", "Cluster_ID",
        "Faturou", "Oportunidade_Real", "Ruido",
        "Cluster_Tem_Faturado", "ID_Ultimo_Cluster",
        "Motivo_Ruido",
        "Valor",
    ]
    debug_clusters = base[debug_cols].copy()

    debug_id_sheet = None
    if args.debug_id is not None:
        base_ids = set(pd.to_numeric(base["ID_Orcamento"], errors="coerce").dropna().astype(int).tolist())
        if int(args.debug_id) in base_ids:
            row = base.loc[
                pd.to_numeric(base["ID_Orcamento"], errors="coerce").astype("Int64") == int(args.debug_id)
            ].iloc[0]
            vend = row["Vendedor"]
            cnpj = row["CNPJ"]
            clid = row["Cluster_ID"]
            debug_id_sheet = base[
                (base["Vendedor"] == vend) & (base["CNPJ"] == cnpj) & (base["Cluster_ID"] == clid)
            ][debug_cols].copy()

    # -----------------------------
    # ExportaÃ§Ã£o
    # -----------------------------
    original_title_map = {
        "Data_Faturamento": "Data de Faturamento",
        "Data_Referencia_WR_Faturamento": "Data Ref. Win Rate Faturamento",
        "Enviado_Aprovacao": col_etapa2,
        "Faturou": col_etapa4,
    }

    def with_selected_original_titles(df_in: pd.DataFrame) -> pd.DataFrame:
        cols = {k: v for k, v in original_title_map.items() if k in df_in.columns}
        return fix_df_headers_mojibake(df_in.rename(columns=cols))

    base_export = with_selected_original_titles(base)
    perdas_reais_resumo_export = with_selected_original_titles(perdas_reais_resumo)
    perdas_reais_definitivas_export = with_selected_original_titles(perdas_reais_definitivas_resumo)
    ruidos_export = with_selected_original_titles(ruidos)
    debug_clusters_export = with_selected_original_titles(debug_clusters)
    debug_id_sheet_export = with_selected_original_titles(debug_id_sheet) if debug_id_sheet is not None else None
    # Mantém nomes internos na aba consumida pelo Script 2.
    oportunidades_export = with_selected_original_titles(oportunidades)
    oportunidades_ref_periodo_export = with_selected_original_titles(oportunidades_ref_periodo)
    pivot_perdas_status_export = fix_df_headers_mojibake(pivot_perdas_status)
    pivot_perdas_status_definitivas_export = fix_df_headers_mojibake(pivot_perdas_status_definitivas)
    comparativo_total_export = fix_df_headers_mojibake(comparativo_total)
    comparativo_total_data_fat_export = fix_df_headers_mojibake(comparativo_total_data_fat)
    config_export = fix_df_headers_mojibake(config)

    output_path = build_output_path(args.output)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        base_export.to_excel(excel_writer=writer, sheet_name="Detalhe_Linha_a_Linha", index=False)
        oportunidades_export.to_excel(excel_writer=writer, sheet_name="Lista_Oportunidades_Reais", index=False)
        oportunidades_ref_periodo_export.to_excel(
            excel_writer=writer,
            sheet_name="Lista_WR_Data_Fat",
            index=False,
        )
        perdas_reais_resumo_export.to_excel(excel_writer=writer, sheet_name="Lista_Nao_Convertidas", index=False)
        pivot_perdas_status_export.to_excel(excel_writer=writer, sheet_name="Pivot_Nao_Convertidas_Status", index=False)
        perdas_reais_definitivas_export.to_excel(excel_writer=writer, sheet_name="Lista_NC_Recusa_Cliente", index=False)
        pivot_perdas_status_definitivas_export.to_excel(
            excel_writer=writer,
            sheet_name="Pivot_NC_Status_Recusa_Cliente",
            index=False,
        )
        ruidos_export.to_excel(excel_writer=writer, sheet_name="Lista_Ruidos", index=False)
        comparativo_total_export.to_excel(excel_writer=writer, sheet_name="Comparativo_Geral_Total", index=False)
        comparativo_total_data_fat_export.to_excel(
            excel_writer=writer,
            sheet_name="Comp_Geral_Total_Data_Fat",
            index=False,
        )
        config_export.to_excel(excel_writer=writer, sheet_name="Config_Modelo", index=False)
        debug_clusters_export.to_excel(excel_writer=writer, sheet_name="Debug_Clusters", index=False)

        if debug_id_sheet_export is not None:
            debug_id_sheet_export.to_excel(excel_writer=writer, sheet_name=f"Debug_ID_{args.debug_id}", index=False)

    aplicar_formatacao_excel(output_path)

    elapsed = time.perf_counter() - start_time
    elapsed_min = int(elapsed // 60)
    elapsed_sec = int(elapsed % 60)

    print(f"Arquivo gerado: {output_path}")
    print(f"Tempo total: {elapsed_min}min {elapsed_sec}s")


if __name__ == "__main__":
    main()


