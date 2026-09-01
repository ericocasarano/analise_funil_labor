#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Acompanha a evolucao diaria, dentro do mes comercial, das vendas de um item
especifico (por CodigoErp): quanto foi Faturado por dia vs. quanto foi
Aprovado pelo Cliente mas ainda nao Faturado (definicao ampla, nao a mesma
regra de "nao convertidos" usada em gerar_itens_perdas_reais.py).

Antes de classificar, remove "ruido" (recotacoes do mesmo negocio) usando a
mesma logica de cluster por Vendedor+CNPJ+tempo+similaridade de itens do
gerar_oportunidades_reais_codes.py, para nao contar o mesmo negocio mais de
uma vez so porque foi reorcado.
"""

import argparse
import glob
import json
import os
import re
import subprocess
import sys
import time
import unicodedata
from datetime import date, datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

DEFAULT_CODIGO = "5401"
CLASSES = ["Faturado", "Aprovados_Cliente_Nao_Faturado", "Enviados_Cliente_Nao_Aprovado"]


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Cruza itens x orcamentos para acompanhar, dia a dia dentro do mes "
            "comercial, quanto de um item foi faturado vs. aprovado pelo cliente "
            "e ainda nao faturado, removendo recotacoes duplicadas (ruido)."
        )
    )
    parser.add_argument(
        "-i",
        "--orcamentos",
        default=None,
        help="Arquivo de orcamentos tratado (.xlsx). Se omitido, usa o mais recente em entrada/.",
    )
    parser.add_argument(
        "-it",
        "--itens",
        default=None,
        help="Arquivo de itens tratado (.xlsx). Se omitido, usa o mais recente em entrada/.",
    )
    parser.add_argument(
        "--codigo",
        default=DEFAULT_CODIGO,
        help=f"Codigo ERP do item a acompanhar. Padrao: {DEFAULT_CODIGO}",
    )
    parser.add_argument(
        "--todos-itens",
        action="store_true",
        help=(
            "Ignora --codigo e gera um resumo (Orcamentos/Quantidade/Valor por categoria) para "
            "todos os codigos de item encontrados no arquivo de itens, num unico relatorio leve."
        ),
    )
    parser.add_argument(
        "--start",
        default=None,
        help="Data de inicio (YYYY-MM-DD). Padrao: inicio do mes comercial atual.",
    )
    parser.add_argument(
        "--end",
        default=None,
        help="Data de fim (YYYY-MM-DD). Padrao: hoje.",
    )
    parser.add_argument(
        "--delta_horas",
        type=float,
        default=360,
        help="Janela em horas para agrupar recotacoes no mesmo cluster (remocao de ruido). Padrao: 360.",
    )
    parser.add_argument(
        "--sim_min",
        type=float,
        default=0.50,
        help="Similaridade minima (Jaccard) de itens para cluster (remocao de ruido). Padrao: 0.50.",
    )
    parser.add_argument(
        "--manter-ruido",
        action="store_true",
        help="Desativa a remocao de ruido (recotacoes) e conta cada orcamento como um caso separado.",
    )
    parser.add_argument(
        "--sem-atualizar-base",
        action="store_true",
        help=(
            "Nao busca a extracao mais recente no OneDrive/SharePoint (paths.json_dir) para tratar; "
            "usa apenas o ultimo arquivo ja tratado em entrada/."
        ),
    )
    parser.add_argument(
        "-o",
        "--output",
        default="evolucao_diaria_item",
        help="Nome base do arquivo de saida (sem timestamp).",
    )
    return parser.parse_args()


def find_latest_entrada_file(prefixo: str) -> str:
    padrao = os.path.join("entrada", f"{prefixo}_*.xlsx")
    candidatos = sorted(glob.glob(padrao))
    if not candidatos:
        raise ValueError(
            f"Nenhum arquivo encontrado com o padrao '{padrao}'. "
            "Rode o fluxo diario antes ou informe o arquivo manualmente (-i/-it)."
        )
    return candidatos[-1]


def load_json_dir_from_config(config_path: str = "automacao_config.json") -> str:
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Arquivo de configuracao nao encontrado: {config_path}")
    with open(config_path, encoding="utf-8") as f:
        cfg = json.load(f)
    json_dir = cfg.get("paths", {}).get("json_dir")
    if not json_dir:
        raise ValueError("Chave 'paths.json_dir' nao encontrada em automacao_config.json.")
    return json_dir


def find_latest_raw_json(json_dir: str, prefixo: str) -> str:
    padrao = os.path.join(json_dir, f"{prefixo}_*.json")
    candidatos = sorted(glob.glob(padrao))
    if not candidatos:
        raise ValueError(f"Nenhum arquivo encontrado com o padrao '{padrao}'.")
    return candidatos[-1]


def garantir_arquivo_tratado(raw_json_path: str, treat_script: str, output_prefix: str) -> str:
    match = re.search(r"(\d{8}_\d{6})", os.path.basename(raw_json_path))
    if not match:
        raise ValueError(f"Nao foi possivel identificar o timestamp no nome do arquivo: {raw_json_path}")
    timestamp = match.group(1)
    output_path = os.path.join("entrada", f"{output_prefix}_{timestamp}.xlsx")
    if os.path.exists(output_path):
        return output_path

    resultado = subprocess.run(
        [sys.executable, treat_script, "-i", raw_json_path, "--timestamp", timestamp],
        capture_output=True,
        text=True,
    )
    if resultado.returncode != 0:
        raise RuntimeError(
            f"Falha ao tratar '{raw_json_path}' com {treat_script}:\n{resultado.stdout}\n{resultado.stderr}"
        )
    return output_path


def clean_text(value):
    if value is None:
        return ""
    return str(value).replace("﻿", "").replace("​", "").strip()


def normalize_key(value):
    text = clean_text(value).lower()
    text = "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip()


def clean_columns(df):
    df = df.copy()
    df.columns = [clean_text(col) for col in df.columns]
    return df


def pick_first_existing(df, candidates):
    normalized = {normalize_key(col): col for col in df.columns}
    for candidate in candidates:
        key = normalize_key(candidate)
        if key in normalized:
            return normalized[key]
    return None


def pick_doc_col(df):
    return pick_first_existing(df, ["CNPJ/ CPF", "CNPJ/CPF", "CNPJ", "Cnpj"])


def normalize_doc_to_14(series: pd.Series) -> pd.Series:
    s = series.astype("string")
    s = s.str.replace(r"\D", "", regex=True)
    s = s.fillna("")
    s = s.str.zfill(14)
    s = s.where(s != "00000000000000", "")
    return s


def jaccard_codes(a: set, b: set) -> float:
    if not a or not b:
        return 0.0
    inter = len(a & b)
    uni = len(a | b)
    return inter / uni if uni > 0 else 0.0


def easter_sunday(year: int) -> date:
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = (h + l - 7 * m + 114) % 31 + 1
    return date(year, month, day)


def brazil_national_holidays(year: int) -> set:
    easter = easter_sunday(year)
    return {
        date(year, 1, 1),
        easter - timedelta(days=2),
        date(year, 4, 21),
        date(year, 5, 1),
        date(year, 9, 7),
        date(year, 10, 12),
        date(year, 11, 2),
        date(year, 11, 15),
        date(year, 11, 20),
        date(year, 12, 25),
    }


def is_business_day(d: date) -> bool:
    if d.weekday() >= 5:
        return False
    return d not in brazil_national_holidays(d.year)


def penultimate_business_day(year: int, month: int) -> date:
    if month == 12:
        next_month = date(year + 1, 1, 1)
    else:
        next_month = date(year, month + 1, 1)
    cursor = next_month - timedelta(days=1)
    business_days = []
    while cursor.month == month:
        if is_business_day(cursor):
            business_days.append(cursor)
            if len(business_days) == 2:
                return business_days[1]
        cursor -= timedelta(days=1)
    raise ValueError(f"Nao foi possivel identificar o penultimo dia util de {month}/{year}.")


def resolve_commercial_period(reference_date: date):
    current_close = penultimate_business_day(reference_date.year, reference_date.month)

    if reference_date <= current_close:
        commercial_end = current_close
        prev_month_date = reference_date.replace(day=1) - timedelta(days=1)
        previous_close = penultimate_business_day(prev_month_date.year, prev_month_date.month)
        commercial_start = previous_close + timedelta(days=1)
    else:
        commercial_start = current_close + timedelta(days=1)
        if reference_date.month == 12:
            next_month_date = date(reference_date.year + 1, 1, 1)
        else:
            next_month_date = date(reference_date.year, reference_date.month + 1, 1)
        commercial_end = penultimate_business_day(next_month_date.year, next_month_date.month)

    return commercial_start, commercial_end


def read_itens(path: str) -> pd.DataFrame:
    df = clean_columns(pd.read_excel(path, dtype={"IDOrcamentoPrinc": "string", "CodigoErp": "string"}))
    required = ["IDOrcamentoPrinc", "CodigoErp"]
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise ValueError(
            "Colunas obrigatorias ausentes no arquivo de itens:\n"
            + "\n".join(f"- {col}" for col in missing)
            + "\n\nColunas encontradas:\n"
            + ", ".join(df.columns.astype(str).tolist())
        )
    df["IDOrcamentoPrinc"] = pd.to_numeric(df["IDOrcamentoPrinc"], errors="coerce").astype("Int64")
    df = df.dropna(subset=["IDOrcamentoPrinc"]).copy()
    df["IDOrcamentoPrinc"] = df["IDOrcamentoPrinc"].astype(int)
    df["CodigoErp"] = df["CodigoErp"].astype("string").fillna("").str.strip()
    df["Quantidade"] = pd.to_numeric(df["Quantidade"], errors="coerce") if "Quantidade" in df.columns else np.nan
    df["VlTot"] = pd.to_numeric(df["VlTot"], errors="coerce") if "VlTot" in df.columns else np.nan
    return df


def read_orcamentos(path: str, precisa_ruido: bool) -> pd.DataFrame:
    df = clean_columns(
        pd.read_excel(
            path,
            dtype={"Cnpj": "string", "CNPJ": "string", "CNPJ/ CPF": "string", "CNPJ/CPF": "string"},
        )
    )

    col_id = pick_first_existing(df, ["Núm. Orç.", "Num. Orc.", "ID_Orcamento"])
    col_data_criacao = pick_first_existing(df, ["Data de Criação", "Data de Criacao", "Data"])
    col_data_fat = pick_first_existing(df, ["Data de Faturamento"])
    col_faturou = pick_first_existing(df, ["Faturou", "ETAPA 4 FUNIL Pedidos Faturados"])
    col_aprovado = pick_first_existing(
        df,
        ["Aprovado pelo Cliente", "Aprovados pelo Cliente", "ETAPA 3 FUNIL Aprovados pelo Cliente"],
    )
    col_status = pick_first_existing(df, ["Status Atual"])
    col_cliente = pick_first_existing(df, ["Cliente"])
    col_tipo_cliente = pick_first_existing(df, ["Tipo Cliente"])
    col_valor = pick_first_existing(df, ["Valor"])
    col_revisao_gestor = pick_first_existing(df, ["Passou por Revisão Gestor", "Passou por Revisao Gestor"])

    required = {
        "Núm. Orç.": col_id,
        "Data de Criação": col_data_criacao,
        "Faturou": col_faturou,
        "Aprovado pelo Cliente": col_aprovado,
        "Valor": col_valor,
    }

    col_enviado = col_vendedor = col_doc = None
    if precisa_ruido:
        col_enviado = pick_first_existing(
            df,
            ["ETAPA 2 FUNIL Orçamentos em Negociação", "ETAPA 2 FUNIL Orcamentos em Negociacao"],
        )
        col_vendedor = pick_first_existing(df, ["Vendedor"])
        col_doc = pick_doc_col(df)
        required["ETAPA 2 FUNIL Orçamentos em Negociação"] = col_enviado
        required["Vendedor"] = col_vendedor
        required["CNPJ/CPF"] = col_doc

    missing = [name for name, col in required.items() if col is None]
    if missing:
        raise ValueError(
            "Colunas obrigatorias ausentes no arquivo de orcamentos:\n"
            + "\n".join(f"- {name}" for name in missing)
            + "\n\nColunas encontradas:\n"
            + ", ".join(df.columns.astype(str).tolist())
        )

    renames = {
        col_id: "Num_Orcamento",
        col_data_criacao: "Data_Criacao",
        col_faturou: "Faturou",
        col_aprovado: "Aprovado_Cliente",
        col_valor: "Valor",
    }
    if col_data_fat:
        renames[col_data_fat] = "Data_Faturamento"
    if col_status:
        renames[col_status] = "Status_Atual"
    if col_cliente:
        renames[col_cliente] = "Cliente"
    if col_tipo_cliente:
        renames[col_tipo_cliente] = "Tipo_Cliente"
    if col_revisao_gestor:
        renames[col_revisao_gestor] = "Revisao_Gestor"
    if precisa_ruido:
        renames[col_enviado] = "Enviado_Aprovacao"
        renames[col_vendedor] = "Vendedor"
        renames[col_doc] = "CNPJ"

    df = df.rename(columns=renames)
    df["Num_Orcamento"] = pd.to_numeric(df["Num_Orcamento"], errors="coerce").astype("Int64")
    df = df.dropna(subset=["Num_Orcamento"]).copy()
    df["Num_Orcamento"] = df["Num_Orcamento"].astype(int)
    df["Data_Criacao"] = pd.to_datetime(df["Data_Criacao"], errors="coerce")
    if "Data_Faturamento" in df.columns:
        df["Data_Faturamento"] = pd.to_datetime(df["Data_Faturamento"], errors="coerce")
    df["Faturou"] = pd.to_numeric(df["Faturou"], errors="coerce").fillna(0).astype(int)
    df["Aprovado_Cliente"] = pd.to_numeric(df["Aprovado_Cliente"], errors="coerce").fillna(0).astype(int)
    df["Valor"] = pd.to_numeric(df["Valor"], errors="coerce")
    if "Status_Atual" in df.columns:
        df["Status_Atual"] = df["Status_Atual"].astype("string").fillna("").str.strip()
    if "Cliente" in df.columns:
        df["Cliente"] = df["Cliente"].astype("string").fillna("").str.strip()
    if "Tipo_Cliente" in df.columns:
        df["Tipo_Cliente"] = df["Tipo_Cliente"].astype("string").fillna("Nao informado").str.strip()
        df.loc[df["Tipo_Cliente"] == "", "Tipo_Cliente"] = "Nao informado"
    if "Revisao_Gestor" in df.columns:
        df["Revisao_Gestor"] = pd.to_numeric(df["Revisao_Gestor"], errors="coerce").fillna(0).astype(int)
    if precisa_ruido:
        df["Enviado_Aprovacao"] = pd.to_numeric(df["Enviado_Aprovacao"], errors="coerce").fillna(0).astype(int)
        df["Vendedor"] = df["Vendedor"].astype("string").fillna("").str.strip()
        df["CNPJ"] = normalize_doc_to_14(df["CNPJ"])
    return df


def remover_ruido(df: pd.DataFrame, mapa_codes: dict, delta_horas: float, sim_min: float):
    """
    Replica a deduplicacao de recotacoes do gerar_oportunidades_reais_codes.py:
    agrupa orcamentos do mesmo Vendedor+CNPJ criados perto no tempo e com itens
    parecidos (Jaccard) num mesmo cluster. Dentro do cluster, so o(s) faturado(s)
    contam como oportunidade real (ou, se nenhum faturou, so o ultimo orcamento
    do cluster) - o restante e "ruido" (recotacao do mesmo negocio).
    """
    base = df[df["Enviado_Aprovacao"] == 1].copy()
    base = base.sort_values(["Vendedor", "CNPJ", "Data_Criacao", "Num_Orcamento"]).reset_index(drop=True)

    base["DeltaHoras"] = base.groupby(["Vendedor", "CNPJ"])["Data_Criacao"].diff().dt.total_seconds() / 3600
    base["Prev_ID"] = base.groupby(["Vendedor", "CNPJ"])["Num_Orcamento"].shift(1)

    sims = []
    for cur, prev in zip(base["Num_Orcamento"].tolist(), base["Prev_ID"].tolist()):
        if pd.isna(prev):
            sims.append(np.nan)
            continue
        sims.append(jaccard_codes(mapa_codes.get(int(cur), set()), mapa_codes.get(int(prev), set())))
    base["Itens_Similarity"] = sims

    base["Crit_Tempo_OK"] = base["DeltaHoras"].notna() & (base["DeltaHoras"] <= delta_horas)
    base["Crit_Itens_OK"] = base["Itens_Similarity"].fillna(0.0) >= sim_min
    base["Relacionado"] = base["Crit_Tempo_OK"] & base["Crit_Itens_OK"]

    base["Novo_Cluster"] = (~base["Relacionado"]).astype(int)
    base.loc[base.groupby(["Vendedor", "CNPJ"]).cumcount() == 0, "Novo_Cluster"] = 1
    base["Cluster_ID"] = base.groupby(["Vendedor", "CNPJ"])["Novo_Cluster"].cumsum()

    base["Oportunidade_Real"] = 0
    base["Ruido"] = 0
    base["Motivo_Ruido"] = ""
    base["Orcamento_Considerado_Cluster"] = ""

    for _, grp in base.groupby(["Vendedor", "CNPJ", "Cluster_ID"], sort=False):
        # Um cluster e cortado em segmentos a cada faturado: tudo que vem ANTES
        # de um faturado, dentro do mesmo segmento, foi superado por ele
        # (recotacao da mesma negociacao) e vira ruido. Um orcamento criado
        # DEPOIS de um faturado inicia um segmento novo - nao e mais a mesma
        # negociacao (que ja fechou), entao e avaliado de forma independente.
        segmento = []
        for idx, faturou in zip(grp.index, grp["Faturou"]):
            segmento.append(idx)
            if faturou == 1:
                referencia = str(grp.loc[idx, "Num_Orcamento"])
                base.loc[idx, "Oportunidade_Real"] = 1
                anteriores = segmento[:-1]
                if anteriores:
                    base.loc[anteriores, "Ruido"] = 1
                    base.loc[anteriores, "Motivo_Ruido"] = "Mesmo cluster com orcamento(s) faturado(s)"
                    base.loc[anteriores, "Orcamento_Considerado_Cluster"] = referencia
                segmento = []

        if segmento:
            last_idx = segmento[-1]
            referencia = str(grp.loc[last_idx, "Num_Orcamento"])
            base.loc[last_idx, "Oportunidade_Real"] = 1
            anteriores = segmento[:-1]
            if anteriores:
                base.loc[anteriores, "Ruido"] = 1
                base.loc[anteriores, "Motivo_Ruido"] = "Sem faturado no cluster; considerado apenas o ultimo orcamento (recotacao mais recente)"
                base.loc[anteriores, "Orcamento_Considerado_Cluster"] = referencia

    oportunidades = base[base["Oportunidade_Real"] == 1].copy()
    ruidos = base[base["Ruido"] == 1].copy()
    return base, oportunidades, ruidos


def join_orcamentos(values) -> str:
    ids = sorted(int(v) for v in values)
    return ", ".join(str(v) for v in ids)


def classificar(row) -> str:
    if row["Faturou"] == 1:
        return "Faturado"
    if row["Aprovado_Cliente"] == 1:
        return "Aprovados_Cliente_Nao_Faturado"
    return "Enviados_Cliente_Nao_Aprovado"


def gerar_resumo_todos_itens(itens: pd.DataFrame, working_df: pd.DataFrame) -> pd.DataFrame:
    itens_por_orc = itens.groupby(["CodigoErp", "IDOrcamentoPrinc"], as_index=False).agg(
        Quantidade_Item=("Quantidade", "sum"),
        Valor_Item=("VlTot", "sum"),
    )
    descricao_por_codigo = (
        itens.groupby("CodigoErp")["Descricao"]
        .agg(lambda s: s.mode().iloc[0] if len(s.mode()) else "")
        .to_dict()
    )

    merged = working_df.merge(
        itens_por_orc, left_on="Num_Orcamento", right_on="IDOrcamentoPrinc", how="inner"
    )
    if merged.empty:
        raise ValueError("Nenhum item foi encontrado nos orcamentos do periodo apos a remocao de ruido.")

    merged["Classificacao"] = merged.apply(classificar, axis=1)

    agg = merged.groupby(["CodigoErp", "Classificacao"], as_index=False).agg(
        Orcamentos=("Num_Orcamento", "nunique"),
        Quantidade=("Quantidade_Item", "sum"),
        Valor=("Valor_Item", "sum"),
        Lista_Orcamentos=("Num_Orcamento", join_orcamentos),
    )

    tabela = agg.pivot_table(
        index="CodigoErp",
        columns="Classificacao",
        values=["Orcamentos", "Quantidade", "Valor"],
        fill_value=0,
    )
    tabela.columns = [f"{metric}_{classe}" for metric, classe in tabela.columns]
    tabela = tabela.reset_index()

    lista_pivot = agg.pivot(index="CodigoErp", columns="Classificacao", values="Lista_Orcamentos")
    lista_pivot.columns = [f"Numeros_Orcamentos_{classe}" for classe in lista_pivot.columns]
    lista_pivot = lista_pivot.fillna("").reset_index()
    tabela = tabela.merge(lista_pivot, on="CodigoErp", how="left")

    for classe in CLASSES:
        for metric in ["Orcamentos", "Quantidade", "Valor"]:
            col = f"{metric}_{classe}"
            if col not in tabela.columns:
                tabela[col] = 0
        lista_col = f"Numeros_Orcamentos_{classe}"
        if lista_col not in tabela.columns:
            tabela[lista_col] = ""
        else:
            tabela[lista_col] = tabela[lista_col].fillna("")

    tabela.insert(1, "Descricao", tabela["CodigoErp"].map(descricao_por_codigo).fillna(""))
    tabela["Total_Orcamentos"] = sum(tabela[f"Orcamentos_{classe}"] for classe in CLASSES)

    decidido = tabela["Orcamentos_Faturado"] + tabela["Orcamentos_Aprovados_Cliente_Nao_Faturado"]
    tabela["Taxa_Conversao_%"] = np.where(
        decidido > 0, (tabela["Orcamentos_Faturado"] / decidido * 100).round(2), np.nan
    )

    ordered_cols = ["CodigoErp", "Descricao", "Total_Orcamentos"]
    for classe in CLASSES:
        ordered_cols += [
            f"Orcamentos_{classe}",
            f"Quantidade_{classe}",
            f"Valor_{classe}",
            f"Numeros_Orcamentos_{classe}",
        ]
    ordered_cols.append("Taxa_Conversao_%")
    tabela = tabela[ordered_cols].sort_values("Valor_Faturado", ascending=False).reset_index(drop=True)
    return tabela


def build_output_path(output_stem: str) -> str:
    historico_dir = os.path.join(os.getcwd(), "historico")
    os.makedirs(historico_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(historico_dir, f"{output_stem}_{timestamp}.xlsx")


def aplicar_formatacao_excel(path_xlsx: str):
    wb = load_workbook(path_xlsx)
    fmt_moeda = "R$ #,##0.00"
    fmt_int = "#,##0"

    for ws in wb.worksheets:
        if ws.max_row < 1:
            continue

        for col in range(1, ws.max_column + 1):
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if not isinstance(header, str):
                continue
            key = normalize_key(header)

            if "valor" in key or "preco" in key:
                fmt = fmt_moeda
            elif "quantidade" in key or "orcamentos" in key:
                fmt = fmt_int
            else:
                continue

            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)) and not (isinstance(cell.value, float) and np.isnan(cell.value)):
                    cell.number_format = fmt

        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                value = ws.cell(row=row, column=col).value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            if max_len > 0:
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 2, 60)

    wb.save(path_xlsx)


def main():
    start_time = time.perf_counter()
    args = parse_args()
    remover_ruido_ativo = not args.manter_ruido

    if (not args.orcamentos or not args.itens) and not args.sem_atualizar_base:
        try:
            json_dir = load_json_dir_from_config()
            if not args.orcamentos:
                raw_orc = find_latest_raw_json(json_dir, "dax_query_base_orcamentos_funil")
                args.orcamentos = garantir_arquivo_tratado(
                    raw_orc, "tratar_dax_orcamentos_json_power_automate.py", "dax_orcamentos_tratado_power_automate"
                )
                print(f"Arquivo de orcamentos (base mais recente do OneDrive, tratada): {args.orcamentos}")
            if not args.itens:
                raw_itens = find_latest_raw_json(json_dir, "dax_query_base_itens_funil")
                args.itens = garantir_arquivo_tratado(
                    raw_itens, "tratar_dax_itens_json_power_automate.py", "dax_itens_tratado_power_automate"
                )
                print(f"Arquivo de itens (base mais recente do OneDrive, tratada): {args.itens}")
        except Exception as exc:
            print(f"Aviso: nao foi possivel atualizar a base a partir do OneDrive ({exc}).")
            print("Usando o ultimo arquivo ja tratado em entrada/.")

    if not args.orcamentos:
        args.orcamentos = find_latest_entrada_file("dax_orcamentos_tratado_power_automate")
        print(f"Arquivo de orcamentos (mais recente local): {args.orcamentos}")
    if not args.itens:
        args.itens = find_latest_entrada_file("dax_itens_tratado_power_automate")
        print(f"Arquivo de itens (mais recente local): {args.itens}")

    itens = read_itens(args.itens)

    mapa_codes = {
        int(oid): set(sub["CodigoErp"].tolist()) for oid, sub in itens.groupby("IDOrcamentoPrinc")
    }

    orcamentos = read_orcamentos(args.orcamentos, precisa_ruido=remover_ruido_ativo)
    orcamentos = orcamentos.dropna(subset=["Data_Criacao"]).copy()

    hoje = date.today()

    if args.start:
        dt_start = pd.to_datetime(args.start, errors="coerce")
        if pd.isna(dt_start):
            raise ValueError(f"Data invalida em --start: '{args.start}'. Use YYYY-MM-DD.")
        commercial_start = dt_start.date()
    else:
        commercial_start, _ = resolve_commercial_period(hoje)

    if args.end:
        dt_end = pd.to_datetime(args.end, errors="coerce")
        if pd.isna(dt_end):
            raise ValueError(f"Data invalida em --end: '{args.end}'. Use YYYY-MM-DD.")
        end_date = dt_end.date()
    else:
        end_date = hoje

    mask = (orcamentos["Data_Criacao"].dt.date >= commercial_start) & (
        orcamentos["Data_Criacao"].dt.date <= end_date
    )
    df_periodo = orcamentos[mask].copy()

    if df_periodo.empty:
        raise ValueError(
            f"Nenhum orcamento criado entre {commercial_start:%d/%m/%Y} e {end_date:%d/%m/%Y}."
        )

    ruidos = None
    if remover_ruido_ativo:
        _, oportunidades, ruidos = remover_ruido(df_periodo, mapa_codes, args.delta_horas, args.sim_min)
        working_df = oportunidades
    else:
        working_df = df_periodo

    if args.todos_itens:
        tabela_todos = gerar_resumo_todos_itens(itens, working_df)

        qtd_ruido_geral = int(ruidos["Num_Orcamento"].nunique()) if ruidos is not None else 0
        valor_ruido_geral = float(ruidos["Valor"].sum()) if ruidos is not None else 0.0

        resumo_geral = pd.DataFrame(
            [
                {"Metrica": "Arquivo de orcamentos", "Valor": args.orcamentos},
                {"Metrica": "Arquivo de itens", "Valor": args.itens},
                {"Metrica": "Periodo - inicio (mes comercial)", "Valor": commercial_start.strftime("%d/%m/%Y")},
                {"Metrica": "Periodo - fim", "Valor": end_date.strftime("%d/%m/%Y")},
                {
                    "Metrica": "Remocao de ruido (recotacoes)",
                    "Valor": "Ativa" if remover_ruido_ativo else "Desativada (--manter-ruido)",
                },
                {"Metrica": "Orcamentos removidos como ruido (todos os itens)", "Valor": qtd_ruido_geral},
                {"Metrica": "Valor removido como ruido (R$, orcamento cheio)", "Valor": round(valor_ruido_geral, 2)},
                {"Metrica": "Codigos de item no relatorio", "Valor": int(tabela_todos["CodigoErp"].nunique())},
                {
                    "Metrica": "Observacao",
                    "Valor": (
                        "Valores de Quantidade/Valor por categoria sao especificos de cada item "
                        "(nao do orcamento inteiro). Ver definicoes de cada categoria rodando o "
                        "script para um --codigo especifico."
                    ),
                },
            ]
        )

        if args.output == "evolucao_diaria_item":
            args.output = "resumo_diario_todos_itens"
        output_path = build_output_path(args.output)
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            resumo_geral.to_excel(writer, sheet_name="Resumo_Geral", index=False)
            tabela_todos.to_excel(writer, sheet_name="Resumo_Por_Item", index=False)
        aplicar_formatacao_excel(output_path)

        elapsed = int(round(time.perf_counter() - start_time))
        mm, ss = divmod(elapsed, 60)
        print(f"Arquivo gerado: {output_path}")
        print(f"Periodo: {commercial_start:%d/%m/%Y} a {end_date:%d/%m/%Y}")
        print(f"Codigos de item no relatorio: {tabela_todos['CodigoErp'].nunique()}")
        print(f"Tempo total: {mm}min {ss}s")
        return

    codigo_alvo = clean_text(args.codigo)
    itens_alvo = itens[itens["CodigoErp"] == codigo_alvo].copy()
    if itens_alvo.empty:
        raise ValueError(f"Nenhum item com CodigoErp = '{codigo_alvo}' foi encontrado no arquivo de itens.")

    itens_agrupados = itens_alvo.groupby("IDOrcamentoPrinc", as_index=False).agg(
        Quantidade_Item=("Quantidade", "sum"),
        Valor_Item=("VlTot", "sum"),
    )

    qtd_ruido_removido = 0
    valor_ruido_removido = 0.0
    ruidos_export = None
    if remover_ruido_ativo:
        ruidos_item = ruidos.merge(
            itens_agrupados, left_on="Num_Orcamento", right_on="IDOrcamentoPrinc", how="inner"
        )
        qtd_ruido_removido = int(ruidos_item["Num_Orcamento"].nunique())
        valor_ruido_removido = float(ruidos_item["Valor_Item"].sum())

        ruido_cols = ["Num_Orcamento"]
        if "Cliente" in ruidos_item.columns:
            ruido_cols.append("Cliente")
        ruido_cols.append("Vendedor")
        ruido_cols.append("Data_Criacao")
        if "Status_Atual" in ruidos_item.columns:
            ruido_cols.append("Status_Atual")
        ruido_cols += [
            "Quantidade_Item",
            "Valor_Item",
            "Motivo_Ruido",
            "Orcamento_Considerado_Cluster",
        ]
        ruidos_export = ruidos_item[ruido_cols].sort_values("Data_Criacao").reset_index(drop=True)

    periodo = working_df.merge(
        itens_agrupados, left_on="Num_Orcamento", right_on="IDOrcamentoPrinc", how="inner"
    ).drop(columns=["IDOrcamentoPrinc"])

    if periodo.empty:
        raise ValueError(
            f"Nenhum orcamento com item '{codigo_alvo}' encontrado entre "
            f"{commercial_start:%d/%m/%Y} e {end_date:%d/%m/%Y} apos a remocao de ruido."
        )

    periodo["Classificacao"] = periodo.apply(classificar, axis=1)
    periodo["Dia"] = periodo["Data_Criacao"].dt.date
    periodo["Valor_Total_Orcamento"] = periodo["Valor"]
    periodo["Preco_Praticado_Item"] = np.where(
        periodo["Quantidade_Item"] > 0, periodo["Valor_Item"] / periodo["Quantidade_Item"], np.nan
    )
    if "Enviado_Aprovacao" in periodo.columns:
        periodo["Enviado_Aprovacao_Label"] = np.where(periodo["Enviado_Aprovacao"] == 1, "Sim", "Nao")
    periodo["Aprovado_Cliente_Label"] = np.where(periodo["Aprovado_Cliente"] == 1, "Sim", "Nao")
    periodo["Faturou_Label"] = np.where(periodo["Faturou"] == 1, "Sim", "Nao")

    evolucao = periodo.groupby(["Dia", "Classificacao"], as_index=False).agg(
        Orcamentos=("Num_Orcamento", "nunique"),
        Quantidade=("Quantidade_Item", "sum"),
        Valor=("Valor_Item", "sum"),
        Lista_Orcamentos=("Num_Orcamento", join_orcamentos),
    )

    tabela = evolucao.pivot_table(
        index="Dia",
        columns="Classificacao",
        values=["Orcamentos", "Quantidade", "Valor"],
        fill_value=0,
    )
    tabela.columns = [f"{metric}_{classe}" for metric, classe in tabela.columns]
    tabela = tabela.reset_index().sort_values("Dia")

    lista_pivot = evolucao.pivot(index="Dia", columns="Classificacao", values="Lista_Orcamentos")
    lista_pivot.columns = [f"Numeros_Orcamentos_{classe}" for classe in lista_pivot.columns]
    lista_pivot = lista_pivot.fillna("").reset_index()
    tabela = tabela.merge(lista_pivot, on="Dia", how="left")

    for classe in CLASSES:
        for metric in ["Orcamentos", "Quantidade", "Valor"]:
            col = f"{metric}_{classe}"
            if col not in tabela.columns:
                tabela[col] = 0
        lista_col = f"Numeros_Orcamentos_{classe}"
        if lista_col not in tabela.columns:
            tabela[lista_col] = ""
        else:
            tabela[lista_col] = tabela[lista_col].fillna("")

    ordered_cols = ["Dia"]
    for classe in CLASSES:
        ordered_cols += [
            f"Orcamentos_{classe}",
            f"Quantidade_{classe}",
            f"Valor_{classe}",
            f"Numeros_Orcamentos_{classe}",
        ]
    tabela = tabela[ordered_cols]

    for classe in CLASSES:
        tabela[f"Quantidade_{classe}_Acumulado"] = tabela[f"Quantidade_{classe}"].cumsum()
        tabela[f"Valor_{classe}_Acumulado"] = tabela[f"Valor_{classe}"].cumsum()

    detalhe_cols = ["Num_Orcamento"]
    if "Cliente" in periodo.columns:
        detalhe_cols.append("Cliente")
    if "Tipo_Cliente" in periodo.columns:
        detalhe_cols.append("Tipo_Cliente")
    detalhe_cols.append("Data_Criacao")
    if "Data_Faturamento" in periodo.columns:
        detalhe_cols.append("Data_Faturamento")
    if "Status_Atual" in periodo.columns:
        detalhe_cols.append("Status_Atual")
    if "Enviado_Aprovacao_Label" in periodo.columns:
        detalhe_cols.append("Enviado_Aprovacao_Label")
    detalhe_cols += [
        "Aprovado_Cliente_Label",
        "Faturou_Label",
        "Classificacao",
        "Quantidade_Item",
        "Preco_Praticado_Item",
        "Valor_Item",
        "Valor_Total_Orcamento",
    ]

    detalhe = periodo[detalhe_cols].sort_values("Data_Criacao").reset_index(drop=True)
    detalhe = detalhe.rename(
        columns={
            "Tipo_Cliente": "Tipo Cliente",
            "Data_Criacao": "Data de Criacao do Orcamento",
            "Enviado_Aprovacao_Label": "Enviado para Aprovacao do Cliente",
            "Aprovado_Cliente_Label": "Aprovado pelo Cliente",
            "Faturou_Label": "Faturou",
            "Classificacao": "Categoria Final",
            "Quantidade_Item": "Volume/Quantidade do Item",
            "Preco_Praticado_Item": "Preco Praticado do Item",
            "Valor_Item": "Valor Total do Item no Orcamento",
            "Valor_Total_Orcamento": "Valor Total do Orcamento",
        }
    )

    status_por_categoria = None
    if "Status_Atual" in periodo.columns:
        status_por_categoria = periodo.groupby(["Classificacao", "Status_Atual"], as_index=False).agg(
            Orcamentos=("Num_Orcamento", "nunique"),
            Quantidade_Item=("Quantidade_Item", "sum"),
            Valor_Item=("Valor_Item", "sum"),
        )
        ordem_categoria = {classe: i for i, classe in enumerate(CLASSES)}
        status_por_categoria["_ordem"] = status_por_categoria["Classificacao"].map(ordem_categoria)
        status_por_categoria = status_por_categoria.sort_values(
            ["_ordem", "Orcamentos"], ascending=[True, False]
        ).drop(columns="_ordem").reset_index(drop=True)
        status_por_categoria = status_por_categoria.rename(
            columns={
                "Classificacao": "Categoria Final",
                "Status_Atual": "Status Atual",
                "Quantidade_Item": "Volume/Quantidade do Item",
                "Valor_Item": "Valor Total do Item no Orcamento",
            }
        )

    total_faturado_valor = float(periodo.loc[periodo["Classificacao"] == "Faturado", "Valor_Item"].sum())
    total_nao_faturado_valor = float(
        periodo.loc[periodo["Classificacao"] == "Aprovados_Cliente_Nao_Faturado", "Valor_Item"].sum()
    )
    total_decidido_valor = total_faturado_valor + total_nao_faturado_valor
    taxa_conversao = (
        round(total_faturado_valor / total_decidido_valor * 100, 2) if total_decidido_valor > 0 else np.nan
    )

    qtd_faturado = int((periodo["Classificacao"] == "Faturado").sum())
    qtd_nao_faturado = int((periodo["Classificacao"] == "Aprovados_Cliente_Nao_Faturado").sum())
    qtd_nao_convertido = int((periodo["Classificacao"] == "Enviados_Cliente_Nao_Aprovado").sum())

    resumo_linhas = [
        {"Metrica": "Codigo do item", "Valor": codigo_alvo},
        {"Metrica": "Arquivo de orcamentos", "Valor": args.orcamentos},
        {"Metrica": "Arquivo de itens", "Valor": args.itens},
        {"Metrica": "Periodo - inicio (mes comercial)", "Valor": commercial_start.strftime("%d/%m/%Y")},
        {"Metrica": "Periodo - fim", "Valor": end_date.strftime("%d/%m/%Y")},
        {"Metrica": "Remocao de ruido (recotacoes)", "Valor": "Ativa" if remover_ruido_ativo else "Desativada (--manter-ruido)"},
    ]
    if remover_ruido_ativo:
        resumo_linhas += [
            {"Metrica": "Orcamentos removidos como ruido (recotacao)", "Valor": qtd_ruido_removido},
            {"Metrica": "Valor removido como ruido (R$)", "Valor": round(valor_ruido_removido, 2)},
        ]
    resumo_linhas += [
        {"Metrica": "Orcamentos no periodo (apos remocao de ruido)", "Valor": int(periodo["Num_Orcamento"].nunique())},
        {"Metrica": "Faturados - Orcamentos", "Valor": qtd_faturado},
        {"Metrica": "Faturados - Valor (R$)", "Valor": round(total_faturado_valor, 2)},
        {"Metrica": "Aprovados e nao faturados - Orcamentos", "Valor": qtd_nao_faturado},
        {"Metrica": "Aprovados e nao faturados - Valor (R$)", "Valor": round(total_nao_faturado_valor, 2)},
        {"Metrica": "Nao convertidos (qualquer status, ainda nao aprovado nem faturado) - Orcamentos", "Valor": qtd_nao_convertido},
        {
            "Metrica": "Taxa de conversao Faturado / (Faturado + Aprovado nao faturado) (%)",
            "Valor": taxa_conversao,
        },
        {
            "Metrica": "Base considerada em TODAS as categorias abaixo",
            "Valor": (
                "So orcamentos que passaram pela ETAPA 2 FUNIL Orcamentos em Negociacao = 1 (de "
                "fato enviados ao cliente) entram na analise - garantido pela remocao de ruido "
                "quando ativa. Orcamentos que nunca saíram da ETAPA 1 (so emitidos, nunca enviados) "
                "ficam fora de todas as 3 categorias."
            ),
        },
        {
            "Metrica": "Definicao de 'Aprovados_Cliente_Nao_Faturado'",
            "Valor": (
                "Aprovado pelo Cliente = 1 (ETAPA 3 - cliente ja aprovou a proposta, um estagio "
                "DEPOIS do envio) e Faturou = 0 (ETAPA 4) - ou seja, o cliente ja disse sim mas o "
                "pedido ainda nao foi faturado. NAO e a mesma regra de 'nao convertidos' usada em "
                "gerar_itens_perdas_reais.py."
            ),
        },
        {
            "Metrica": "Definicao de 'Enviados_Cliente_Nao_Aprovado'",
            "Valor": (
                "Aprovado pelo Cliente = 0 e Faturou = 0 - o orcamento foi enviado ao cliente mas "
                "ele ainda nao aprovou (nem foi faturado), qualquer que seja o Status Atual "
                "(cancelado, em confeccao, analise de credito, aguardando aprovacao, etc.)."
            ),
        },
        {
            "Metrica": "Definicao de 'ruido' (recotacao)",
            "Valor": (
                "Mesma logica do gerar_oportunidades_reais_codes.py: orcamentos do mesmo Vendedor+CNPJ "
                "criados ate --delta_horas de diferenca e com itens similares (Jaccard >= --sim_min) "
                "formam um cluster; so o(s) faturado(s) do cluster (ou, se nenhum faturou, o ultimo "
                "orcamento) conta - o resto e descartado como recotacao do mesmo negocio. So considera "
                "orcamentos com ETAPA 2 FUNIL Orcamentos em Negociacao = 1."
            ),
        },
    ]
    resumo = pd.DataFrame(resumo_linhas)

    output_path = build_output_path(args.output)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        resumo.to_excel(writer, sheet_name="Resumo", index=False)
        tabela.to_excel(writer, sheet_name="Evolucao_Diaria", index=False)
        detalhe.to_excel(writer, sheet_name="Detalhe_Orcamentos", index=False)
        if status_por_categoria is not None:
            status_por_categoria.to_excel(writer, sheet_name="Status_Por_Categoria", index=False)
        if ruidos_export is not None:
            ruidos_export.to_excel(writer, sheet_name="Ruidos_Removidos", index=False)

    aplicar_formatacao_excel(output_path)

    elapsed = int(round(time.perf_counter() - start_time))
    mm, ss = divmod(elapsed, 60)
    print(f"Arquivo gerado: {output_path}")
    print(f"Item: {codigo_alvo} | Periodo: {commercial_start:%d/%m/%Y} a {end_date:%d/%m/%Y}")
    if remover_ruido_ativo:
        print(f"Ruido removido (recotacoes): {qtd_ruido_removido} orcamentos (R$ {valor_ruido_removido:,.2f})")
    print(f"Orcamentos no periodo: {periodo['Num_Orcamento'].nunique()}")
    print(f"Faturados: {qtd_faturado} (R$ {total_faturado_valor:,.2f})")
    print(f"Aprovados e nao faturados: {qtd_nao_faturado} (R$ {total_nao_faturado_valor:,.2f})")
    print(f"Nao convertidos: {qtd_nao_convertido}")
    print(f"Tempo total: {mm}min {ss}s")


if __name__ == "__main__":
    main()
