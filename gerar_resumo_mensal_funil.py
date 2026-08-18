#!/usr/bin/env python
# -*- coding: utf-8 -*-

import argparse
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from gerar_oportunidades_reais_codes import (
    clean_columns,
    jaccard_codes,
    normalize_doc_to_14,
    normalize_valor_to_numeric,
    pick_doc_col,
    pick_first_existing,
    read_input,
    read_itens,
)
from tratar_dax1_json_power_automate import convert_power_automate_json as convert_dax1_json
from tratar_dax2_json_power_automate import convert_power_automate_json as convert_dax2_json


MONTH_LABELS_PT = {
    1: "Jan",
    2: "Fev",
    3: "Mar",
    4: "Abr",
    5: "Mai",
    6: "Jun",
    7: "Jul",
    8: "Ago",
    9: "Set",
    10: "Out",
    11: "Nov",
    12: "Dez",
}

OUTPUT_COLUMNS = [
    "Mês/Ano Comercial",
    "Periodo",
    "Enviados_Qtd (Sem Ruído)",
    "Enviados_Valor (Sem Ruído)",
    "Faturado_Qtd (Sem Ruído)",
    "Faturado_Valor (Sem Ruído)",
    "Nao_Convertidas_Qtd (Sem Ruído)",
    "Nao_Convertidas_Valor (Sem Ruído)",
    "Win Rate (Volume) % (Sem Ruído)",
    "Win Rate (Valor) % (Sem Ruído)",
]

OUTPUT_SHEET_NAME_CRIACAO = "Resumo_Mensal_Funil"
OUTPUT_SHEET_NAME_DATA_FAT = "Resumo_Mensal_Funil_Data_Fat"


@dataclass(frozen=True)
class CommercialMonth:
    reference_month_start: date
    commercial_start: date
    commercial_end: date
    analyzed_start: date
    analyzed_end: date
    complete_period: bool

    @property
    def month_label(self) -> str:
        return f"{MONTH_LABELS_PT[self.reference_month_start.month]}/{str(self.reference_month_start.year)[-2:]}"

    @property
    def period_label(self) -> str:
        return f"{self.analyzed_start.strftime('%d/%m')} a {self.analyzed_end.strftime('%d/%m')}"

    @property
    def commercial_period_label(self) -> str:
        return f"{self.commercial_start.strftime('%d/%m/%Y')} a {self.commercial_end.strftime('%d/%m/%Y')}"


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Gera um resumo mensal do funil usando a mesma logica de remocao "
            "de ruidos do gerar_oportunidades_reais_codes.py."
        )
    )
    parser.add_argument(
        "-i",
        "--input",
        required=True,
        help="Base DAX1 tratada (.xlsx/.xls/.csv) ou JSON bruto do Power Automate.",
    )
    parser.add_argument(
        "-it",
        "--itens",
        required=True,
        help="Base DAX2 tratada (.xlsx/.xls/.csv) ou JSON bruto do Power Automate.",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="resumo_mensal_funil",
        help="Nome base do arquivo de saida (sem timestamp).",
    )
    parser.add_argument(
        "--delta_horas",
        type=float,
        default=360,
        help="Janela em horas para agrupar no mesmo cluster.",
    )
    parser.add_argument(
        "--sim_min",
        type=float,
        default=0.50,
        help="Similaridade minima (Jaccard) para cluster (0 a 1).",
    )
    parser.add_argument(
        "--modo",
        choices=["comercial", "calendario"],
        default="comercial",
        help="Define se o agrupamento mensal sera por mes comercial ou por mes calendario.",
    )
    parser.add_argument(
        "--start",
        default="",
        help="Data inicial opcional para limitar a base analisada, no formato YYYY-MM-DD.",
    )
    parser.add_argument(
        "--end",
        default="",
        help="Data final opcional para limitar a base analisada, no formato YYYY-MM-DD.",
    )
    return parser.parse_args()


def get_easter_sunday(year: int) -> date:
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def get_brazil_national_holidays(year: int) -> set[date]:
    easter = get_easter_sunday(year)
    return {
        date(year, 1, 1),
        easter - timedelta(days=2),
        date(year, 4, 21),
        date(year, 5, 1),
        date(year, 9, 7),
        date(year, 10, 12),
        date(year, 11, 2),
        date(year, 11, 15),
        date(year, 11, 20),
        date(year, 12, 25),
    }


def is_business_day(current_date: date) -> bool:
    if current_date.weekday() >= 5:
        return False
    return current_date not in get_brazil_national_holidays(current_date.year)


def shift_month(any_date: date, months: int) -> date:
    month_index = (any_date.year * 12 + any_date.month - 1) + months
    year = month_index // 12
    month = month_index % 12 + 1
    return date(year, month, 1)


def last_day_of_month(month_start: date) -> date:
    return shift_month(month_start, 1) - timedelta(days=1)


def get_penultimate_business_day_of_month(year: int, month: int) -> date:
    cursor = last_day_of_month(date(year, month, 1))
    business_days: list[date] = []

    while cursor.month == month:
        if is_business_day(cursor):
            business_days.append(cursor)
            if len(business_days) == 2:
                return business_days[1]
        cursor -= timedelta(days=1)

    raise ValueError(f"Nao foi possivel identificar o penultimo dia util de {month:02d}/{year}.")


def build_output_path(output_stem: str) -> Path:
    historico_dir = Path.cwd() / "historico"
    historico_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return historico_dir / f"{output_stem}_{timestamp}.xlsx"


def load_dax1_dataframe(path_str: str) -> pd.DataFrame:
    path = Path(path_str).resolve()
    ext = path.suffix.lower()
    if ext == ".json":
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_output = Path(tmp_dir) / "dax1_tratado.xlsx"
            return convert_dax1_json(path, temp_output)
    return clean_columns(read_input(str(path)))


def load_dax2_dataframe(path_str: str) -> pd.DataFrame:
    path = Path(path_str).resolve()
    ext = path.suffix.lower()
    if ext == ".json":
        with tempfile.TemporaryDirectory() as tmp_dir:
            temp_output = Path(tmp_dir) / "dax2_tratado.xlsx"
            return convert_dax2_json(path, temp_output)
    return clean_columns(read_itens(str(path)))


def prepare_base_dataframe(df_raw: pd.DataFrame) -> pd.DataFrame:
    col_doc = pick_doc_col(df_raw.columns)
    if not col_doc:
        raise ValueError("Coluna de documento nao encontrada. Esperado: Cnpj / CNPJ / CNPJ/ CPF / CNPJ/CPF")

    col_id_orc = pick_first_existing(df_raw, ["Núm. Orç.", "Num. Orc.", "NÃºm. OrÃ§."])
    col_data_criacao = pick_first_existing(df_raw, ["Data de Criação", "Data de Criacao", "Data de CriaÃ§Ã£o"])
    col_data_faturamento = pick_first_existing(
        df_raw,
        [
            "Data de Faturamento",
            "Data Faturamento",
            "Data_Faturamento",
            "Dt Faturamento",
            "Dt_Faturamento",
        ],
    )
    col_etapa2 = pick_first_existing(
        df_raw,
        [
            "ETAPA 2 FUNIL Orçamentos em Negociação",
            "ETAPA 2 FUNIL Orcamentos em Negociacao",
            "ETAPA 2 FUNIL OrÃ§amentos em NegociaÃ§Ã£o",
        ],
    )
    col_etapa4 = pick_first_existing(
        df_raw,
        [
            "ETAPA 4 FUNIL Pedidos Faturados",
            "ETAPA 4 FUNIL Pedidos faturados",
        ],
    )

    required_base = {
        "ID_Orcamento": col_id_orc,
        "Data": col_data_criacao,
        "CNPJ": col_doc,
        "Vendedor": "Vendedor" if "Vendedor" in df_raw.columns else None,
        "Valor": "Valor" if "Valor" in df_raw.columns else None,
        "Enviado_Aprovacao": col_etapa2,
        "Faturou": col_etapa4,
    }
    missing = [key for key, value in required_base.items() if value is None]
    if missing:
        raise ValueError(
            "Colunas obrigatorias ausentes na base DAX1:\n"
            + "\n".join(f"- {col}" for col in missing)
            + "\n\nColunas encontradas:\n"
            + ", ".join(df_raw.columns.astype(str).tolist())
        )

    df = df_raw.rename(
        columns={
            col_id_orc: "ID_Orcamento",
            col_data_criacao: "Data",
            col_doc: "CNPJ",
            "Vendedor": "Vendedor",
            "Valor": "Valor",
            col_etapa2: "Enviado_Aprovacao",
            col_etapa4: "Faturou",
        }
    ).copy()
    if col_data_faturamento:
        df = df.rename(columns={col_data_faturamento: "Data_Faturamento"})

    df["Data"] = pd.to_datetime(df["Data"], errors="coerce")
    if "Data_Faturamento" not in df.columns:
        df["Data_Faturamento"] = pd.NaT
    df["Data_Faturamento"] = pd.to_datetime(df["Data_Faturamento"], errors="coerce")
    df["Valor"] = normalize_valor_to_numeric(df["Valor"])
    df["Faturou"] = pd.to_numeric(df["Faturou"], errors="coerce").fillna(0).astype(int)
    df["Enviado_Aprovacao"] = pd.to_numeric(df["Enviado_Aprovacao"], errors="coerce").fillna(0).astype(int)
    df["CNPJ"] = normalize_doc_to_14(df["CNPJ"])
    df["Data_Referencia_Faturamento"] = df["Data_Faturamento"].where(
        df["Data_Faturamento"].notna(),
        df["Data"],
    )

    df = df.dropna(subset=["Data"]).copy()
    df["ID_Orcamento"] = pd.to_numeric(df["ID_Orcamento"], errors="coerce").astype("Int64")
    df = df.dropna(subset=["ID_Orcamento"]).copy()
    df["ID_Orcamento"] = df["ID_Orcamento"].astype(int)
    return df


def build_items_map(itens_raw: pd.DataFrame) -> dict[int, set[str]]:
    required_itens = ["IDOrcamentoPrinc", "CodigoErp"]
    missing = [column for column in required_itens if column not in itens_raw.columns]
    if missing:
        raise ValueError(
            "Colunas obrigatorias ausentes na base DAX2:\n"
            + "\n".join(f"- {col}" for col in missing)
            + "\n\nColunas encontradas:\n"
            + ", ".join(itens_raw.columns.astype(str).tolist())
        )

    itens = itens_raw.copy()
    itens["IDOrcamentoPrinc"] = pd.to_numeric(itens["IDOrcamentoPrinc"], errors="coerce").astype("Int64")
    itens["CodigoErp"] = itens["CodigoErp"].astype("string").fillna("").str.strip()
    itens = itens.dropna(subset=["IDOrcamentoPrinc"])
    itens = itens[itens["IDOrcamentoPrinc"] > 0]

    return {
        int(orcamento_id): set(sub_df["CodigoErp"].tolist())
        for orcamento_id, sub_df in itens.groupby("IDOrcamentoPrinc")
    }


def build_commercial_months(min_date: date, max_date: date) -> list[CommercialMonth]:
    commercial_months: list[CommercialMonth] = []
    cursor = date(min_date.year, min_date.month, 1)
    last_reference_month = date(max_date.year, max_date.month, 1)

    while cursor <= last_reference_month:
        previous_month = shift_month(cursor, -1)
        commercial_start = get_penultimate_business_day_of_month(previous_month.year, previous_month.month) + timedelta(days=1)
        commercial_end = get_penultimate_business_day_of_month(cursor.year, cursor.month)
        analyzed_start = max(commercial_start, min_date)
        analyzed_end = min(commercial_end, max_date)

        if analyzed_start <= analyzed_end:
            commercial_months.append(
                CommercialMonth(
                    reference_month_start=cursor,
                    commercial_start=commercial_start,
                    commercial_end=commercial_end,
                    analyzed_start=analyzed_start,
                    analyzed_end=analyzed_end,
                    complete_period=(analyzed_start == commercial_start and analyzed_end == commercial_end),
                )
            )

        cursor = shift_month(cursor, 1)

    return commercial_months


def build_calendar_months(min_date: date, max_date: date) -> list[CommercialMonth]:
    calendar_months: list[CommercialMonth] = []
    cursor = date(min_date.year, min_date.month, 1)
    last_reference_month = date(max_date.year, max_date.month, 1)

    while cursor <= last_reference_month:
        calendar_start = cursor
        calendar_end = last_day_of_month(cursor)
        analyzed_start = max(calendar_start, min_date)
        analyzed_end = min(calendar_end, max_date)

        if analyzed_start <= analyzed_end:
            calendar_months.append(
                CommercialMonth(
                    reference_month_start=cursor,
                    commercial_start=calendar_start,
                    commercial_end=calendar_end,
                    analyzed_start=analyzed_start,
                    analyzed_end=analyzed_end,
                    complete_period=(analyzed_start == calendar_start and analyzed_end == calendar_end),
                )
            )

        cursor = shift_month(cursor, 1)

    return calendar_months


def apply_date_filters(df: pd.DataFrame, start_str: str, end_str: str) -> pd.DataFrame:
    filtered = df.copy()

    if start_str:
        start_dt = pd.to_datetime(start_str, errors="coerce")
        if pd.isna(start_dt):
            raise ValueError(f"Data inicial invalida: {start_str}. Use o formato YYYY-MM-DD.")
        filtered = filtered[filtered["Data"] >= start_dt].copy()

    if end_str:
        end_dt = pd.to_datetime(end_str, errors="coerce")
        if pd.isna(end_dt):
            raise ValueError(f"Data final invalida: {end_str}. Use o formato YYYY-MM-DD.")
        filtered = filtered[filtered["Data"] <= end_dt].copy()

    return filtered


def apply_broad_date_filters(df: pd.DataFrame, start_str: str, end_str: str) -> pd.DataFrame:
    filtered = df.copy()
    reference_col = "Data_Referencia_Faturamento"

    if start_str:
        start_dt = pd.to_datetime(start_str, errors="coerce")
        if pd.isna(start_dt):
            raise ValueError(f"Data inicial invalida: {start_str}. Use o formato YYYY-MM-DD.")
        filtered = filtered[
            (filtered["Data"] >= start_dt)
            | (filtered[reference_col] >= start_dt)
        ].copy()

    if end_str:
        end_dt = pd.to_datetime(end_str, errors="coerce")
        if pd.isna(end_dt):
            raise ValueError(f"Data final invalida: {end_str}. Use o formato YYYY-MM-DD.")
        filtered = filtered[
            (filtered["Data"] <= end_dt)
            | (filtered[reference_col] <= end_dt)
        ].copy()

    return filtered


def build_clustered_opportunities(
    base_df: pd.DataFrame,
    items_map: dict[int, set[str]],
    delta_horas: float,
    sim_min: float,
) -> pd.DataFrame:
    base = base_df[base_df["Enviado_Aprovacao"] == 1].copy()

    if base.empty:
        return base

    base = base.sort_values(["Vendedor", "CNPJ", "Data", "ID_Orcamento"]).reset_index(drop=True)
    base["DeltaHoras"] = base.groupby(["Vendedor", "CNPJ"])["Data"].diff().dt.total_seconds() / 3600
    base["Prev_ID"] = base.groupby(["Vendedor", "CNPJ"])["ID_Orcamento"].shift(1)

    similarities: list[float] = []
    for current_id, previous_id in zip(base["ID_Orcamento"].tolist(), base["Prev_ID"].tolist()):
        if pd.isna(previous_id):
            similarities.append(np.nan)
            continue

        similarities.append(
            jaccard_codes(
                items_map.get(int(current_id), set()),
                items_map.get(int(previous_id), set()),
            )
        )

    base["Itens_Similarity"] = similarities
    base["Crit_Tempo_OK"] = base["DeltaHoras"].notna() & (base["DeltaHoras"] <= delta_horas)
    base["Crit_Itens_OK"] = base["Itens_Similarity"].fillna(0.0) >= sim_min
    base["Relacionado"] = base["Crit_Tempo_OK"] & base["Crit_Itens_OK"]
    base["Novo_Cluster"] = (~base["Relacionado"]).astype(int)
    base.loc[base.groupby(["Vendedor", "CNPJ"]).cumcount() == 0, "Novo_Cluster"] = 1
    base["Cluster_ID"] = base.groupby(["Vendedor", "CNPJ"])["Novo_Cluster"].cumsum()

    base["Oportunidade_Real"] = 0
    for (_, _, _), group_df in base.groupby(["Vendedor", "CNPJ", "Cluster_ID"], sort=False):
        faturados_indexes = group_df.index[group_df["Faturou"] == 1].tolist()

        if faturados_indexes:
            base.loc[faturados_indexes, "Oportunidade_Real"] = 1
        else:
            last_index = group_df.index[-1]
            base.loc[last_index, "Oportunidade_Real"] = 1

    oportunidades = base[base["Oportunidade_Real"] == 1].copy()
    oportunidades["Valor_Faturado"] = np.where(oportunidades["Faturou"] == 1, oportunidades["Valor"], 0.0)
    return oportunidades


def summarize_period(
    oportunidades_df: pd.DataFrame,
    period_start: date,
    period_end: date,
    date_col: str,
) -> dict[str, float]:
    start_ts = pd.Timestamp(period_start)
    end_ts = pd.Timestamp(period_end)

    oportunidades = oportunidades_df[
        (oportunidades_df[date_col] >= start_ts)
        & (oportunidades_df[date_col] <= end_ts)
    ].copy()

    if oportunidades.empty:
        return {
            "Enviados_Qtd (Sem Ruído)": 0,
            "Enviados_Valor (Sem Ruído)": 0.0,
            "Faturado_Qtd (Sem Ruído)": 0,
            "Faturado_Valor (Sem Ruído)": 0.0,
            "Nao_Convertidas_Qtd (Sem Ruído)": 0,
            "Nao_Convertidas_Valor (Sem Ruído)": 0.0,
            "Win Rate (Volume) % (Sem Ruído)": 0.0,
            "Win Rate (Valor) % (Sem Ruído)": 0.0,
        }

    enviados_qtd = int(oportunidades["ID_Orcamento"].count())
    faturado_qtd = int(oportunidades["Faturou"].sum())
    nao_convertidas_qtd = enviados_qtd - faturado_qtd

    enviados_valor = float(oportunidades["Valor"].sum(skipna=True))
    faturado_valor = float(oportunidades["Valor_Faturado"].sum(skipna=True))
    nao_convertidas_valor = enviados_valor - faturado_valor

    wr_volume = (faturado_qtd / enviados_qtd * 100) if enviados_qtd > 0 else 0.0
    wr_valor = (faturado_valor / enviados_valor * 100) if enviados_valor > 0 else 0.0

    return {
        "Enviados_Qtd (Sem Ruído)": enviados_qtd,
        "Enviados_Valor (Sem Ruído)": enviados_valor,
        "Faturado_Qtd (Sem Ruído)": faturado_qtd,
        "Faturado_Valor (Sem Ruído)": faturado_valor,
        "Nao_Convertidas_Qtd (Sem Ruído)": nao_convertidas_qtd,
        "Nao_Convertidas_Valor (Sem Ruído)": nao_convertidas_valor,
        "Win Rate (Volume) % (Sem Ruído)": wr_volume,
        "Win Rate (Valor) % (Sem Ruído)": wr_valor,
    }


def summarize_period_by_creation(
    base_df: pd.DataFrame,
    items_map: dict[int, set[str]],
    period_start: date,
    period_end: date,
    delta_horas: float,
    sim_min: float,
) -> dict[str, float]:
    start_ts = pd.Timestamp(period_start)
    end_ts = pd.Timestamp(period_end)
    base_periodo = base_df[
        (base_df["Data"] >= start_ts)
        & (base_df["Data"] <= end_ts)
    ].copy()

    oportunidades_periodo = build_clustered_opportunities(
        base_df=base_periodo,
        items_map=items_map,
        delta_horas=delta_horas,
        sim_min=sim_min,
    )

    return summarize_period(
        oportunidades_df=oportunidades_periodo,
        period_start=period_start,
        period_end=period_end,
        date_col="Data",
    )


def format_output(output_path: Path):
    workbook = load_workbook(output_path)
    target_sheets = [OUTPUT_SHEET_NAME_CRIACAO, OUTPUT_SHEET_NAME_DATA_FAT]

    currency_columns = {"D", "F", "H"}
    percent_columns = {"I", "J"}

    for sheet_name in target_sheets:
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]

        for column_cells in worksheet.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                cell.alignment = Alignment(vertical="center")
                if cell.row > 1 and column_letter in currency_columns:
                    cell.number_format = 'R$ #,##0.00'
                if cell.row > 1 and column_letter in percent_columns:
                    cell.number_format = '0.00%'
                value_length = len(str(cell.value)) if cell.value is not None else 0
                max_length = max(max_length, value_length)

            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 28)

        worksheet.freeze_panes = "A2"
    workbook.save(output_path)


def main():
    args = parse_args()

    dax1_raw = load_dax1_dataframe(args.input)
    dax2_raw = load_dax2_dataframe(args.itens)

    base_df_full = prepare_base_dataframe(dax1_raw)
    base_df_criacao = apply_date_filters(base_df_full, args.start, args.end)
    base_df_broad = apply_broad_date_filters(base_df_full, args.start, args.end)
    if base_df_criacao.empty:
        raise ValueError("A base DAX1 ficou vazia depois do filtro informado.")

    items_map = build_items_map(dax2_raw)
    oportunidades_df = build_clustered_opportunities(
        base_df=base_df_broad,
        items_map=items_map,
        delta_horas=args.delta_horas,
        sim_min=args.sim_min,
    )
    if oportunidades_df.empty:
        raise ValueError("Nenhuma oportunidade real foi identificada na base informada.")

    if args.start:
        min_date = pd.to_datetime(args.start, errors="coerce").date()
    else:
        min_date = base_df_criacao["Data"].min().date()
    if args.end:
        max_date = pd.to_datetime(args.end, errors="coerce").date()
    else:
        max_date = max(
            base_df_criacao["Data"].max().date(),
            oportunidades_df["Data_Referencia_Faturamento"].dropna().max().date(),
        )
    if args.modo == "comercial":
        analysis_months = build_commercial_months(min_date, max_date)
    else:
        analysis_months = build_calendar_months(min_date, max_date)

    if not analysis_months:
        raise ValueError("Nenhum mes foi identificado dentro da base informada.")

    rows_criacao: list[dict] = []
    rows_data_fat: list[dict] = []
    partial_months: list[CommercialMonth] = []

    for analysis_month in analysis_months:
        metrics_criacao = summarize_period_by_creation(
            base_df=base_df_criacao,
            items_map=items_map,
            period_start=analysis_month.analyzed_start,
            period_end=analysis_month.analyzed_end,
            delta_horas=args.delta_horas,
            sim_min=args.sim_min,
        )
        metrics_data_fat = summarize_period(
            oportunidades_df=oportunidades_df,
            period_start=analysis_month.analyzed_start,
            period_end=analysis_month.analyzed_end,
            date_col="Data_Referencia_Faturamento",
        )

        base_row = {
            "Mês/Ano Comercial": analysis_month.month_label,
            "Periodo": analysis_month.period_label,
        }
        rows_criacao.append(
            {
                **base_row,
                **metrics_criacao,
            }
        )
        rows_data_fat.append(
            {
                **base_row,
                **metrics_data_fat,
            }
        )

        if not analysis_month.complete_period:
            partial_months.append(analysis_month)

    def finalize_output(rows: list[dict]) -> pd.DataFrame:
        output_df = pd.DataFrame(rows)[OUTPUT_COLUMNS]
        for column in [
            "Enviados_Valor (Sem Ruído)",
            "Faturado_Valor (Sem Ruído)",
            "Nao_Convertidas_Valor (Sem Ruído)",
        ]:
            output_df[column] = pd.to_numeric(output_df[column], errors="coerce").fillna(0.0)

        for column in [
            "Win Rate (Volume) % (Sem Ruído)",
            "Win Rate (Valor) % (Sem Ruído)",
        ]:
            output_df[column] = pd.to_numeric(output_df[column], errors="coerce").fillna(0.0) / 100.0
        return output_df

    output_df_criacao = finalize_output(rows_criacao)
    output_df_data_fat = finalize_output(rows_data_fat)

    output_path = build_output_path(args.output)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        output_df_criacao.to_excel(writer, sheet_name=OUTPUT_SHEET_NAME_CRIACAO, index=False)
        output_df_data_fat.to_excel(writer, sheet_name=OUTPUT_SHEET_NAME_DATA_FAT, index=False)

    format_output(output_path)

    print(f"Arquivo gerado: {output_path}")
    print(f"Linhas geradas (criacao): {len(output_df_criacao)}")
    print(f"Linhas geradas (data faturamento): {len(output_df_data_fat)}")
    print(f"Base analisada: {min_date.strftime('%d/%m/%Y')} a {max_date.strftime('%d/%m/%Y')}")
    print(f"Modo de agrupamento: {args.modo}")

    if partial_months:
        print("Meses com periodo parcial por limite da base:")
        for item in partial_months:
            print(
                f"- {item.month_label}: analisado {item.period_label} | periodo cheio {item.commercial_period_label}"
            )


if __name__ == "__main__":
    main()
