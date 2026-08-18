#!/usr/bin/env python
# -*- coding: utf-8 -*-

import argparse
import os
import re
import time
import unicodedata
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


DEFAULT_TARGET_CODES = ["5401", "8007", "581"]


def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Analisa orçamentos faturados da aba Lista_Oportunidades_Reais "
            "que contêm códigos-alvo e ranqueia itens coocorrentes."
        )
    )
    parser.add_argument(
        "-i",
        "--input",
        required=True,
        help="Arquivo gerado pelo gerar_oportunidades_reais_codes.py (.xlsx).",
    )
    parser.add_argument(
        "-it",
        "--itens",
        required=True,
        help="Arquivo itens_do_orcamento (.xlsx/.xls/.csv).",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="analise_itens_codigos_alvo",
        help="Nome base do arquivo de saída (sem timestamp).",
    )
    parser.add_argument(
        "--codes",
        nargs="+",
        default=DEFAULT_TARGET_CODES,
        help="Lista de códigos-alvo. Ex.: --codes 5401 8007 581",
    )
    parser.add_argument(
        "--sheet",
        default="Lista_Oportunidades_Reais",
        help="Nome da aba de oportunidades reais no arquivo de entrada.",
    )
    parser.add_argument(
        "--top",
        type=int,
        default=100,
        help="Quantidade máxima de linhas nas abas de ranking principal.",
    )
    parser.add_argument(
        "--start",
        default=None,
        help="Data início (YYYY-MM-DD) para filtrar a Data do orçamento.",
    )
    parser.add_argument(
        "--end",
        default=None,
        help="Data fim (YYYY-MM-DD) para filtrar a Data do orçamento.",
    )
    return parser.parse_args()


def clean_text(value):
    if value is None:
        return ""
    return str(value).replace("\ufeff", "").replace("\u200b", "").strip()


def normalize_key(value: str) -> str:
    text = clean_text(value).lower()
    text = "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip()


def clean_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [clean_text(col) for col in df.columns]
    return df


def pick_first_existing(df: pd.DataFrame, candidates: list[str]) -> str | None:
    normalized = {normalize_key(col): col for col in df.columns}
    for candidate in candidates:
        key = normalize_key(candidate)
        if key in normalized:
            return normalized[key]
    return None


def read_excel_sheet(path: str, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet_name)
    return clean_columns(df)


def read_itens(path: str) -> pd.DataFrame:
    ext = os.path.splitext(path)[1].lower()
    kwargs = {
        "dtype": {
            "IDOrcamentoPrinc": "string",
            "CodigoErp": "string",
            "Descricao": "string",
        }
    }
    if ext in [".xlsx", ".xls"]:
        return clean_columns(pd.read_excel(path, **kwargs))
    if ext == ".csv":
        return clean_columns(pd.read_csv(path, sep=None, engine="python", encoding="utf-8-sig", **kwargs))
    raise ValueError(f"Extensão não suportada: {ext}. Use .xlsx/.xls/.csv")


def build_output_path(output_stem: str) -> str:
    historico_dir = os.path.join(os.getcwd(), "historico")
    os.makedirs(historico_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(historico_dir, f"{output_stem}_{timestamp}.xlsx")


def normalize_codigo(series: pd.Series) -> pd.Series:
    return series.astype("string").fillna("").str.strip()


def normalize_status(series: pd.Series) -> pd.Series:
    return series.astype("string").fillna("").map(normalize_key)


def aplicar_formatacao_excel(path_xlsx: str):
    wb = load_workbook(path_xlsx)
    fmt_moeda = 'R$ #,##0.00'
    fmt_pct = '0.00"%"'
    fmt_int = "#,##0"
    fmt_float = "#,##0.00"

    for ws in wb.worksheets:
        if ws.max_row < 1:
            continue

        for col in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col)
            header_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if not isinstance(header, str):
                continue
            key = normalize_key(header)

            if "share" in key:
                fmt = fmt_pct
            elif "valor" in key:
                fmt = fmt_moeda
            elif "indice afinidade" in key:
                fmt = fmt_float
            elif (
                "orcamentos" in key
                or "linhas" in key
                or "qtd" in key
                or "quantidade" in key
                or "total" in key
            ):
                fmt = fmt_int
            else:
                continue

            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)) and not (isinstance(cell.value, float) and np.isnan(cell.value)):
                    cell.number_format = fmt

        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, str):
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        for col in range(1, ws.max_column + 1):
            max_len = 0
            for row in range(1, ws.max_row + 1):
                value = ws.cell(row=row, column=col).value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            if max_len > 0:
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_len + 2, 60)

    wb.save(path_xlsx)


def preparar_oportunidades(df: pd.DataFrame) -> pd.DataFrame:
    col_id = pick_first_existing(df, ["ID_Orcamento", "Núm. Orç.", "Num. Orc.", "NÃºm. OrÃ§.", "NÃƒÂºm. OrÃƒÂ§."])
    col_data = pick_first_existing(df, ["Data", "Data de Criação", "Data de Criacao", "Data de CriaÃ§Ã£o"])
    col_status = pick_first_existing(
        df,
        [
            "Status Atual",
            "Status Atual do Orçamento",
            "Status Atual do Orcamento",
            "Status Atual orçamento",
            "Status Atual orcamento",
        ],
    )
    col_faturou = pick_first_existing(df, ["Faturou", "ETAPA 4 FUNIL Pedidos Faturados"])

    required = {
        "ID_Orcamento": col_id,
        "Data": col_data,
        "Status Atual": col_status,
    }
    missing = [name for name, col in required.items() if col is None]
    if missing:
        raise ValueError(
            "Colunas obrigatórias ausentes na aba de oportunidades:\n"
            + "\n".join(f"- {name}" for name in missing)
            + "\n\nColunas encontradas:\n"
            + ", ".join(df.columns.astype(str).tolist())
        )

    renames = {
        col_id: "ID_Orcamento",
        col_data: "Data",
        col_status: "Status Atual",
    }
    if col_faturou:
        renames[col_faturou] = "Faturou"

    out = df.rename(columns=renames).copy()
    out["ID_Orcamento"] = pd.to_numeric(out["ID_Orcamento"], errors="coerce").astype("Int64")
    out = out.dropna(subset=["ID_Orcamento"]).copy()
    out["ID_Orcamento"] = out["ID_Orcamento"].astype(int)
    out["Data"] = pd.to_datetime(out["Data"], errors="coerce")
    out["Status Atual"] = out["Status Atual"].astype("string").fillna("").str.strip()
    out["_status_norm"] = normalize_status(out["Status Atual"])
    if "Tipo Cliente" in out.columns:
        out["Tipo Cliente"] = out["Tipo Cliente"].astype("string").fillna("Não informado").str.strip()
        out.loc[out["Tipo Cliente"] == "", "Tipo Cliente"] = "Não informado"
    else:
        out["Tipo Cliente"] = "Não informado"

    if "Faturou" in out.columns:
        out["Faturou"] = pd.to_numeric(out["Faturou"], errors="coerce").fillna(0).astype(int)
    else:
        out["Faturou"] = 0

    return out


def preparar_itens(df: pd.DataFrame) -> pd.DataFrame:
    required = ["IDOrcamentoPrinc", "CodigoErp", "Descricao"]
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise ValueError(
            "Colunas obrigatórias ausentes no arquivo de itens:\n"
            + "\n".join(f"- {col}" for col in missing)
            + "\n\nColunas encontradas:\n"
            + ", ".join(df.columns.astype(str).tolist())
        )

    out = df.copy()
    out["IDOrcamentoPrinc"] = pd.to_numeric(out["IDOrcamentoPrinc"], errors="coerce").astype("Int64")
    out = out.dropna(subset=["IDOrcamentoPrinc"]).copy()
    out["IDOrcamentoPrinc"] = out["IDOrcamentoPrinc"].astype(int)
    out["CodigoErp"] = normalize_codigo(out["CodigoErp"])
    out["Descricao"] = out["Descricao"].astype("string").fillna("").str.strip()
    out["Quantidade"] = pd.to_numeric(out["Quantidade"], errors="coerce") if "Quantidade" in out.columns else np.nan
    out["VlTot"] = pd.to_numeric(out["VlTot"], errors="coerce") if "VlTot" in out.columns else np.nan
    return out


def rankear_itens(
    itens_grupo: pd.DataFrame,
    itens_base: pd.DataFrame,
    total_orcamentos_grupo: int,
    total_orcamentos_base: int,
    target_codes: set[str],
    top_n: int,
    excluir_codigos_alvo: bool,
) -> pd.DataFrame:
    grupo = itens_grupo.copy()
    base = itens_base.copy()

    if excluir_codigos_alvo:
        grupo = grupo[~grupo["CodigoErp"].isin(target_codes)].copy()
        base = base[~base["CodigoErp"].isin(target_codes)].copy()

    ranking_grupo = (
        grupo.groupby(["CodigoErp", "Descricao"], dropna=False)
        .agg(
            Orcamentos_Com_Item=("IDOrcamentoPrinc", pd.Series.nunique),
            Linhas_Item=("IDOrcamentoPrinc", "count"),
            Quantidade_Total=("Quantidade", "sum"),
            Valor_Total_Item=("VlTot", "sum"),
        )
        .reset_index()
    )

    ranking_base = (
        base.groupby(["CodigoErp", "Descricao"], dropna=False)
        .agg(
            Orcamentos_Base_Faturados=("IDOrcamentoPrinc", pd.Series.nunique),
        )
        .reset_index()
    )

    ranking = ranking_grupo.merge(ranking_base, on=["CodigoErp", "Descricao"], how="left")
    ranking["Orcamentos_Base_Faturados"] = ranking["Orcamentos_Base_Faturados"].fillna(0).astype(int)
    ranking["Share_Orcamentos_Grupo_%"] = np.where(
        total_orcamentos_grupo > 0,
        ranking["Orcamentos_Com_Item"] / total_orcamentos_grupo * 100,
        np.nan,
    )
    ranking["Share_Orcamentos_Base_%"] = np.where(
        total_orcamentos_base > 0,
        ranking["Orcamentos_Base_Faturados"] / total_orcamentos_base * 100,
        np.nan,
    )
    ranking["Indice_Afinidade"] = np.where(
        ranking["Share_Orcamentos_Base_%"] > 0,
        ranking["Share_Orcamentos_Grupo_%"] / ranking["Share_Orcamentos_Base_%"],
        np.nan,
    )

    ranking = ranking.sort_values(
        ["Share_Orcamentos_Grupo_%", "Orcamentos_Com_Item", "Indice_Afinidade", "Valor_Total_Item"],
        ascending=[False, False, False, False],
    ).reset_index(drop=True)

    if top_n > 0:
        ranking = ranking.head(top_n).copy()

    return ranking


def rankear_itens_por_tipo_cliente(
    itens_grupo: pd.DataFrame,
    itens_base: pd.DataFrame,
    target_codes: set[str],
) -> pd.DataFrame:
    grupo = itens_grupo[~itens_grupo["CodigoErp"].isin(target_codes)].copy()
    base = itens_base[~itens_base["CodigoErp"].isin(target_codes)].copy()

    totais_grupo = (
        grupo.groupby("Tipo Cliente", dropna=False)["IDOrcamentoPrinc"]
        .nunique()
        .rename("Total_Orcamentos_Grupo_Tipo")
        .reset_index()
    )
    totais_base = (
        base.groupby("Tipo Cliente", dropna=False)["IDOrcamentoPrinc"]
        .nunique()
        .rename("Total_Orcamentos_Base_Tipo")
        .reset_index()
    )

    ranking_grupo = (
        grupo.groupby(["Tipo Cliente", "CodigoErp", "Descricao"], dropna=False)
        .agg(
            Orcamentos_Com_Item=("IDOrcamentoPrinc", pd.Series.nunique),
            Linhas_Item=("IDOrcamentoPrinc", "count"),
            Quantidade_Total=("Quantidade", "sum"),
            Valor_Total_Item=("VlTot", "sum"),
        )
        .reset_index()
    )
    ranking_base = (
        base.groupby(["Tipo Cliente", "CodigoErp", "Descricao"], dropna=False)
        .agg(
            Orcamentos_Base_Tipo=("IDOrcamentoPrinc", pd.Series.nunique),
        )
        .reset_index()
    )

    ranking = ranking_grupo.merge(ranking_base, on=["Tipo Cliente", "CodigoErp", "Descricao"], how="left")
    ranking = ranking.merge(totais_grupo, on="Tipo Cliente", how="left")
    ranking = ranking.merge(totais_base, on="Tipo Cliente", how="left")

    ranking["Orcamentos_Base_Tipo"] = ranking["Orcamentos_Base_Tipo"].fillna(0).astype(int)
    ranking["Share_Orcamentos_Grupo_Tipo_%"] = np.where(
        ranking["Total_Orcamentos_Grupo_Tipo"] > 0,
        ranking["Orcamentos_Com_Item"] / ranking["Total_Orcamentos_Grupo_Tipo"] * 100,
        np.nan,
    )
    ranking["Share_Orcamentos_Base_Tipo_%"] = np.where(
        ranking["Total_Orcamentos_Base_Tipo"] > 0,
        ranking["Orcamentos_Base_Tipo"] / ranking["Total_Orcamentos_Base_Tipo"] * 100,
        np.nan,
    )
    ranking["Indice_Afinidade_Tipo"] = np.where(
        ranking["Share_Orcamentos_Base_Tipo_%"] > 0,
        ranking["Share_Orcamentos_Grupo_Tipo_%"] / ranking["Share_Orcamentos_Base_Tipo_%"],
        np.nan,
    )

    ranking = ranking.sort_values(
        ["Tipo Cliente", "Share_Orcamentos_Grupo_Tipo_%", "Orcamentos_Com_Item", "Indice_Afinidade_Tipo"],
        ascending=[True, False, False, False],
    ).reset_index(drop=True)

    return ranking


def main():
    start_time = time.perf_counter()
    args = parse_args()

    oportunidades = preparar_oportunidades(read_excel_sheet(args.input, args.sheet))
    itens = preparar_itens(read_itens(args.itens))

    if args.start:
        dt_start = pd.to_datetime(args.start, errors="coerce")
        if pd.isna(dt_start):
            raise ValueError(f"Data inválida em --start: '{args.start}'. Use YYYY-MM-DD.")
        oportunidades = oportunidades[oportunidades["Data"] >= dt_start].copy()

    if args.end:
        dt_end = pd.to_datetime(args.end, errors="coerce")
        if pd.isna(dt_end):
            raise ValueError(f"Data inválida em --end: '{args.end}'. Use YYYY-MM-DD.")
        oportunidades = oportunidades[oportunidades["Data"] < (dt_end + pd.Timedelta(days=1))].copy()

    if oportunidades.empty:
        raise ValueError("Nenhum orçamento encontrado após aplicar o filtro de datas.")

    target_codes = {str(code).strip() for code in args.codes if str(code).strip()}
    if not target_codes:
        raise ValueError("Nenhum código-alvo válido foi informado em --codes.")

    faturados = oportunidades[oportunidades["_status_norm"] == "faturado"].copy()

    if faturados.empty:
        raise ValueError("Nenhum orçamento com Status Atual = Faturado foi encontrado na aba Lista_Oportunidades_Reais.")

    itens_faturados = itens[itens["IDOrcamentoPrinc"].isin(faturados["ID_Orcamento"])].copy()
    if itens_faturados.empty:
        raise ValueError("Nenhum item encontrado para os orçamentos faturados.")

    target_items = itens_faturados[itens_faturados["CodigoErp"].isin(target_codes)].copy()
    target_budget_ids = set(target_items["IDOrcamentoPrinc"].dropna().astype(int).tolist())
    if not target_budget_ids:
        raise ValueError("Nenhum orçamento faturado contém os códigos-alvo informados.")

    orcamentos_alvo = faturados[faturados["ID_Orcamento"].isin(target_budget_ids)].copy()
    itens_orcamentos_alvo = itens_faturados[itens_faturados["IDOrcamentoPrinc"].isin(target_budget_ids)].copy()

    codigos_por_orcamento = (
        target_items.groupby("IDOrcamentoPrinc")["CodigoErp"]
        .apply(lambda s: " | ".join(sorted(set(s.astype(str)))))
        .rename("Codigos_Alvo_Encontrados")
        .reset_index()
    )
    orcamentos_alvo = orcamentos_alvo.merge(
        codigos_por_orcamento,
        left_on="ID_Orcamento",
        right_on="IDOrcamentoPrinc",
        how="left",
    ).drop(columns=["IDOrcamentoPrinc"])

    tipo_cliente_por_orc = faturados[["ID_Orcamento", "Tipo Cliente"]].drop_duplicates()
    itens_faturados = itens_faturados.merge(
        tipo_cliente_por_orc,
        left_on="IDOrcamentoPrinc",
        right_on="ID_Orcamento",
        how="left",
    ).drop(columns=["ID_Orcamento"])
    itens_faturados["Tipo Cliente"] = itens_faturados["Tipo Cliente"].fillna("Não informado")

    itens_orcamentos_alvo = itens_faturados[itens_faturados["IDOrcamentoPrinc"].isin(target_budget_ids)].copy()

    total_faturados = int(faturados["ID_Orcamento"].nunique())
    total_alvo = int(orcamentos_alvo["ID_Orcamento"].nunique())

    ranking_sem_alvo = rankear_itens(
        itens_grupo=itens_orcamentos_alvo,
        itens_base=itens_faturados,
        total_orcamentos_grupo=total_alvo,
        total_orcamentos_base=total_faturados,
        target_codes=target_codes,
        top_n=args.top,
        excluir_codigos_alvo=True,
    )
    ranking_completo = rankear_itens(
        itens_grupo=itens_orcamentos_alvo,
        itens_base=itens_faturados,
        total_orcamentos_grupo=total_alvo,
        total_orcamentos_base=total_faturados,
        target_codes=target_codes,
        top_n=args.top,
        excluir_codigos_alvo=False,
    )
    ranking_por_tipo = rankear_itens_por_tipo_cliente(
        itens_grupo=itens_orcamentos_alvo,
        itens_base=itens_faturados,
        target_codes=target_codes,
    )

    resumo_tipo = (
        orcamentos_alvo.groupby("Tipo Cliente", dropna=False)["ID_Orcamento"]
        .nunique()
        .rename("Orcamentos_Alvo")
        .reset_index()
    )
    resumo_tipo["Total_Faturados_Tipo"] = resumo_tipo["Tipo Cliente"].map(
        faturados.groupby("Tipo Cliente")["ID_Orcamento"].nunique()
    )
    resumo_tipo["Share_Alvo_no_Tipo_%"] = np.where(
        resumo_tipo["Total_Faturados_Tipo"] > 0,
        resumo_tipo["Orcamentos_Alvo"] / resumo_tipo["Total_Faturados_Tipo"] * 100,
        np.nan,
    )
    resumo_tipo = resumo_tipo.sort_values("Orcamentos_Alvo", ascending=False).reset_index(drop=True)

    resumo = pd.DataFrame(
        [
            {"Metrica": "Arquivo de entrada", "Valor": args.input},
            {"Metrica": "Arquivo de itens", "Valor": args.itens},
            {"Metrica": "Aba analisada", "Valor": args.sheet},
            {"Metrica": "Códigos-alvo", "Valor": ", ".join(sorted(target_codes))},
            {"Metrica": "Data início filtro", "Valor": args.start or ""},
            {"Metrica": "Data fim filtro", "Valor": args.end or ""},
            {"Metrica": "Total de orçamentos faturados", "Valor": total_faturados},
            {"Metrica": "Orçamentos faturados com código-alvo", "Valor": total_alvo},
            {
                "Metrica": "Share dos orçamentos-alvo sobre faturados (%)",
                "Valor": round(total_alvo / total_faturados * 100, 2) if total_faturados > 0 else np.nan,
            },
            {
                "Metrica": "Leitura sugerida",
                "Valor": "Priorizar Share_Orcamentos_Grupo_% + Orcamentos_Com_Item e usar Indice_Afinidade para desempatar.",
            },
        ]
    )

    output_path = build_output_path(args.output)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        resumo.to_excel(writer, sheet_name="Resumo", index=False)
        orcamentos_alvo.to_excel(writer, sheet_name="Orcamentos_Alvo", index=False)
        target_items.to_excel(writer, sheet_name="Itens_Codigos_Alvo", index=False)
        itens_orcamentos_alvo.to_excel(writer, sheet_name="Itens_Grupo_Alvo", index=False)
        resumo_tipo.to_excel(writer, sheet_name="Resumo_Tipo_Cliente", index=False)
        ranking_sem_alvo.to_excel(writer, sheet_name="Ranking_Coocorrencia", index=False)
        ranking_completo.to_excel(writer, sheet_name="Ranking_Com_Alvos", index=False)
        ranking_por_tipo.to_excel(writer, sheet_name="Ranking_Por_Tipo", index=False)

    aplicar_formatacao_excel(output_path)

    elapsed = int(round(time.perf_counter() - start_time))
    mm, ss = divmod(elapsed, 60)
    print(f"Arquivo gerado: {output_path}")
    print(f"Orçamentos faturados analisados: {total_faturados}")
    print(f"Orçamentos faturados com código-alvo: {total_alvo}")
    print(f"Tempo total: {mm}min {ss}s")


if __name__ == "__main__":
    main()
